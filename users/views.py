from django.shortcuts import get_object_or_404, render, redirect # Keep this import
from django.contrib.sessions.models import Session
from django.http import HttpResponse, JsonResponse
import openpyxl
import uuid
import json
import io
import os
from django.utils.safestring import mark_safe
from django.contrib.auth import logout, get_user_model # Import get_user_model for User
from django.contrib.auth.decorators import login_required
from django.urls import reverse # Se añade para manejo de redirecciones con parámetros
from django.contrib import messages # Keep this import
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from django.db.models import Q, Count, Case, When, IntegerField, Sum, F # Keep these imports
from paginaPetrocentro.forms import RegisterForm

# Explicit imports from users.models
from .models import Empleado, Project, Tarea, Curso, EmpleadoCurso, LogActividad, MensajeChat, InventoryCategory, Proveedor, InventoryItem, StockTransaction, TransferRequest, Ubicacion, Cargo, Candidato

# Explicit imports from configuracion.models
from configuracion.models import Rol, Rol_permiso, Permisos

# Import Usuario from paginaPetrocentro.models as it's a profile model
from paginaPetrocentro.models import Usuario
from django.utils import timezone
from django.utils.text import slugify
from openpyxl.styles import Font, Alignment
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from django.template.loader import get_template
from django.db import transaction
from xhtml2pdf import pisa

# Get Django's User model
User = get_user_model()

# --- UTILIDADES ---
def recibir_transferencia(request, request_id):
    transferencia = Transferencia.objects.filter(id=request_id).first()
    if not transferencia:
        return JsonResponse({"error": "Transferencia no encontrada"}, status=404)

    transferencia.estado = "RECIBIDA"
    transferencia.save()
    return JsonResponse({"ok": True, "mensaje": "Transferencia recibida"})

def obtener_permisos(permisos):
    crear = 0
    eliminar = 0
    editar = 0
    consultar = 0
    usuarios = 0 
    inventario = 0
   
    for permiso in permisos:
        nombre = permiso.permiso.nombre.lower() # Convertir a minúsculas para comparar
        if nombre == "crear": crear = 1
        elif nombre in ["consultar", "consultor"]: consultar = 1
        elif nombre == "eliminar": eliminar = 1
        elif nombre in ["actualizar", "editar", "editor"]: editar = 1 
        elif nombre == "usuarios": usuarios = 1
        elif nombre == "inventario": inventario = 1
    return {
        'crear': crear, 'consultar': consultar, 'editar': editar, 'eliminar': eliminar,
        'usuarios': usuarios, 'inventario': inventario,
    }

# --- UTILIDADES ---
def obtener_contexto_usuario(request):
    """
    Centraliza la obtención del perfil de usuario, empleado y permisos
    para evitar repetición de código en las vistas.
    """
    usuario_id = request.session.get('usuario_logeado')
    user = request.user
    
    permisos = {}
    # Búsqueda más robusta del perfil de usuario
    usuario_profile = Usuario.objects.filter(id=usuario_id).first()
    if not usuario_profile and user.is_authenticated:
        usuario_profile = Usuario.objects.filter(user_id=user).first()

    if not usuario_profile:
        return {
            'usuario': None, 'empleado': None, 'permisos': {}, 
            'nombre_rol': "Invitado", 'usuarios': 0, 'inventario': 0
        }

    empleado_profile = None
    nombre_rol = "Usuario"

    if user.is_superuser:
        permisos = {'crear': 1, 'consultar': 1, 'editar': 1, 'eliminar': 1, 'usuarios': 1, 'inventario': 1}
        nombre_rol = "ADMINISTRADOR"
        empleado_profile = Empleado.objects.filter(id=usuario_profile.id).select_related('id_rol').first()
    else:
        try:
            empleado_profile = Empleado.objects.get(id=usuario_profile.id)
            if empleado_profile.id_rol:
                nombre_rol = empleado_profile.id_rol.nombre.strip().upper() # Normalizado
                permisos_qs = Rol_permiso.objects.filter(rol=empleado_profile.id_rol)
                permisos = obtener_permisos(permisos_qs)
            else:
                nombre_rol = "Empleado (Sin Rol)"
        except Empleado.DoesNotExist:
            pass
            
    # Normalización de banderas de visibilidad (Sincronizado con context_processors.py)
    is_admin = nombre_rol == "ADMINISTRADOR" or user.is_superuser
    is_supervisor = nombre_rol == "SUPERVISOR" or user.is_superuser
    has_profile = empleado_profile is not None

    if is_admin:
        dashboard_home = 'perfil_admin'
    else:
        dashboard_home = 'perfil_empleado'

    return {
        'usuario': usuario_profile,
        'empleado': empleado_profile,
        'permisos': permisos,
        'nombre_rol': nombre_rol,
        'dashboard_home': dashboard_home,
        # Banderas de Visibilidad del Menú (Sincronización Total con context_processors.py)
        'menu_inicio_visible': True,
        'menu_dashboard_visible': is_admin or is_supervisor,
        'menu_usuarios_visible': is_admin,
        'menu_empleados_visible': is_admin,
        'menu_roles_visible': is_admin,
        'menu_inventario_visible': is_admin or is_supervisor or permisos.get('inventario', 0) == 1,
        'menu_tareas_visible': is_admin or is_supervisor,
        'menu_agenda_visible': has_profile,
        'menu_cursos_visible': is_admin,
        'menu_publicaciones_visible': is_admin,
        'menu_auditoria_visible': is_admin,
        'menu_rrhh_visible': is_admin,
        'menu_pqrs_visible': is_admin,
        'menu_cotizaciones_visible': is_admin,
        'menu_nosotros_visible': is_admin,
        'menu_perfil_visible': True,
        # Banderas de compatibilidad con código antiguo
        'usuarios': permisos.get('usuarios', 0) == 1 or is_admin,
        'inventario': 1 if (permisos.get('inventario', 0) == 1 or is_admin or is_supervisor) else 0,
        'crear': permisos.get('crear', 0) == 1,
        'consultar': permisos.get('consultar', 0) == 1,
        'editar': permisos.get('editar', 0) == 1,
        'eliminar': permisos.get('eliminar', 0) == 1,
        'is_supervisor': is_supervisor,
        'is_admin': is_admin,
    }

#------------------------------------------------------EMPLEADOS Y USUARIOS----------------------------------------------
@login_required
def listar_empleados(request):
    ctx = obtener_contexto_usuario(request)
    if not ctx['empleado'] and not request.user.is_superuser:
        messages.error(request, 'No tienes un perfil de empleado para acceder a esta página.')
        return redirect('index')

    permisos = ctx['permisos']
 
    # LÓGICA DE ACCESO:
    # Si tiene permiso 'usuarios' (Admin/Gestor), ve todos los empleados y puede registrar nuevos.
    if permisos.get('usuarios') == 1:
            lista_empleados = Empleado.objects.select_related('user_id', 'id_rol', 'id_cargo', 'id_ubicacion').all()
            users_para_registro = Usuario.objects.all()
    else:
            # Si es empleado regular, entra a la ruta pero SOLO ve su propia información.
            lista_empleados = Empleado.objects.select_related('user_id', 'id_rol', 'id_cargo', 'id_ubicacion').filter(id=ctx['usuario'].id)
            users_para_registro = Usuario.objects.none() # No puede registrar a otros
        
    #----------------------------------------------------------------------------------------------------------------
    # FILTROS
    busqueda = request.GET.get("buscar") 
    fecha = request.GET.get("fechaI")
    estado= request.GET.get("estado")
    filtro_cargo = request.GET.get("cargo")
    filtro_ubicacion = request.GET.get("ubicacion")
 
    if estado and estado != "0":
        lista_empleados = lista_empleados.filter(Q(estado = estado)).distinct()
        
    if busqueda:
        lista_empleados = lista_empleados.filter(
            Q(id__icontains = busqueda)|
            Q(nombre__icontains = busqueda)|
            Q(telefono__icontains= busqueda)|
            Q(identificacion__icontains= busqueda)|
            Q(correo__icontains = busqueda)|
            Q(id_rol__nombre__icontains = busqueda)|
            Q(id_cargo__nombre__icontains = busqueda)|
            Q(id_ubicacion__nombre__icontains = busqueda)
        )
 
    if fecha: 
        lista_empleados = lista_empleados.filter(Q(fecha_ingreso__icontains = fecha)).distinct()
 
    if filtro_cargo and filtro_cargo != '0':
        lista_empleados = lista_empleados.filter(id_cargo_id=filtro_cargo)
    
    if filtro_ubicacion and filtro_ubicacion != '0':
        lista_empleados = lista_empleados.filter(id_ubicacion_id=filtro_ubicacion)

    #----------------------------------------------------------------------------------------------------------------   

    # Optimizamos la consulta con anotaciones para evitar consultas dentro del bucle
    p = Paginator(lista_empleados.annotate(total_cursos_count=Count('cursos_asignados', distinct=True), aprobados_count=Count('cursos_asignados', filter=Q(cursos_asignados__estado='APROBADO'), distinct=True), total_tareas_count=Count('tarea', distinct=True), completadas_count=Count('tarea', filter=Q(tarea__completada=True), distinct=True)), 5)
    page_number = request.GET.get('page')
    pagina= p.get_page(page_number)

    # Calcular nivel de cumplimiento para los empleados mostrados
    for emp in pagina:
        emp.cumplimiento = int((emp.aprobados_count / emp.total_cursos_count) * 100) if emp.total_cursos_count > 0 else 0
        emp.cumplimiento_tareas = int((emp.completadas_count / emp.total_tareas_count) * 100) if emp.total_tareas_count > 0 else 0

        
    data ={
        'cargos': Cargo.objects.all(),
        **ctx, # Unpack the common context
        'cursos': Curso.objects.all(), # Para el modal de asignar curso
        'ubicaciones': Ubicacion.objects.all(),
        'roles': Rol.objects.all(),
        'users': users_para_registro, # Lista de usuarios para el modal de registro                  
        'paginas': pagina,
        'paginator': p,
        'filtros': {'buscar': busqueda, 'estado': estado, 'cargo': filtro_cargo, 'ubicacion': filtro_ubicacion, 'fecha': fecha}
        }
        
    return render(request,'dash/empleados.html',data)

    #Obtener todos las tablas de la base de datos

@login_required
def editar_empleados(request, id):
    ctx = obtener_contexto_usuario(request)
 
    # --- Verificación de permisos para EDITAR ---    
    if not ctx['editar']:
        messages.error(request, 'No tienes permisos para editar empleados.')
        return redirect('empleados')

    if request.method == 'POST':
        #consulta para obtener el empleado que coincida con el id que envia el botón de editar
        usuario = Empleado.objects.get(id=id)
        

        #----------------------------------------------------------------------------------------------------------------
        try:    
            #Variable para almacenar el cargo enviado en el formulario, ya que solo se permite editar el cargo
            id_cargo = request.POST.get('cargo')
        
            id_rol = request.POST.get('rol')
            telefono = request.POST.get('telefono1')
            identificacion = request.POST.get('identificacion1')
            #En la variable usuario, se obtiene el cargo y se realiza una consulta que trae el cargo de la tabla "CARGO" para validar coincidencias y asi mismo cambiarlo en el empleado
            usuario.telefono = telefono
            usuario.identificacion = identificacion
           
            usuario.id_cargo = Cargo.objects.get(id_cargo = id_cargo)
        
            if id_rol != '#':
                usuario.id_rol = Rol.objects.get(id_rol = id_rol)
            else:
                usuario.id_rol = None
            #gurdar el cargo
            usuario.save()
            
            # --- LOG DE ACTIVIDAD ---
            try:
                admin_resp = ctx['usuario']
                LogActividad.objects.create(
                    usuario_afectado=usuario, # Empleado hereda de Usuario
                    admin_responsable=admin_resp,
                    accion="Edición de Empleado",
                    detalles=f"Se actualizaron datos (Cargo/Rol/Teléfono) de {usuario.nombre}"
                )
            except:
                pass
            
            #si es exitoso el cambio, envía un mensaje de éxito.
            if(usuario):
                messages.success(request, f'Empleado {usuario.nombre} actualizado correctamente.')
            
          
            #si no es exitoso, envía un mensaje de error
            else:
                messages.error(request, 'No se ha editado exitosamente')
                
        #excepciones para validar errores
        except (KeyError, ValueError) as e:
            messages.error(request, e)
            
    return redirect('empleados')

@login_required
def registrar_empleados(request):
    ctx = obtener_contexto_usuario(request)

    # --- Verificación de permisos para CREAR ---    
    if not ctx['crear']:
        messages.error(request, 'No tienes permisos para registrar empleados.')
        return redirect('empleados')

    if request.method == 'POST':
        
        #----------------------------------------------------------------------------------------------------------------
        #try para pedirle que en caso que falle, arroje el error en un mensaje y no el error de django 

            #Obtener toda la informacion del dormulario de empleados
            fecha = request.POST.get("fechaIngreso")
            id_cargo = request.POST.get('cargo')
            id_rol= request.POST.get('rol')
            identificacion = request.POST.get('identificacion')
            telefono = request.POST.get('telefono')
            ubicacion = request.POST.get('ubicacion')
            usuario_id= request.POST.get('usuario_id')
            
            
            #----------------------------------------------------------------------------------------------------------------
            #Obtener las fechas y convertirlas 
            fechaH = datetime.strptime(fecha, '%Y-%m-%d').date()
        
            #Obtener fecha de hoy 
            fecha_hoy= datetime.today().date()
            
            #----------------------------------------------------------------------------------------------------------------
            #Obtener el id_user que tiene el usuario
            user = Usuario.objects.get(id = usuario_id)
            usuario_num = user.user_id.pk

            #----------------------------------------------------------------------------------------------------------------

            if not id_rol or id_rol == "#":
            #instancia de empleado para crear el mismo
                empleado= Empleado(
                
                        id=user.id,  # Usar la clave primaria del usuario
                        user_id= User.objects.get(id=usuario_num),  #la clave de user_auth para el usuario
                        identificacion= identificacion,
                        estado=user.estado,
                        nombre=user.nombre,
                        correo=user.correo,
                        telefono=telefono,
                        fecha_ingreso=fecha,
                        
                        id_cargo= Cargo.objects.get(id_cargo = id_cargo),
                        
  
                        id_ubicacion=Ubicacion.objects.get(idUbicacion = ubicacion),

                        )
                
            else:
                        empleado= Empleado(
                
                        id=user.id,  # Usar la clave primaria del usuario
                        user_id= User.objects.get(id=usuario_num),  #la clave de user_auth para el usuario
                        identificacion= identificacion,
                        estado=user.estado,
                        nombre=user.nombre,
                        correo=user.correo,
                        telefono=telefono,
                        fecha_ingreso=fecha,
                        id_rol = Rol.objects.get(id_rol = id_rol ),
                        id_cargo= Cargo.objects.get(id_cargo = id_cargo),
                    
                        id_ubicacion=Ubicacion.objects.get(idUbicacion = ubicacion),

                        )
            #----------------------------------------------------------------------------------------------------------------
            #Condición para valdiar si el empleado ya existe
            if (Empleado.objects.filter(id=usuario_id)).exists():
                messages.error(request, 'No se ha registrado exitosamente, el empleado ya existe.')
            
            #----------------------------------------------------------------------------------------------------------------            
            #condicion para validar que el empleado no se puede registrar un dia mayor al actual
            elif(fechaH > fecha_hoy):
                messages.error(request, 'No se puede registrar una fecha superior al dia presente.')
            
            #----------------------------------------------------------------------------------------------------------------            
            #si todo lo anterior se cumple, guarde el empleado
            else:
                empleado.save()
                messages.success(request, f'Empleado {empleado.nombre} registrado exitosamente.')
                

    return redirect('empleados')

#funcion para cambiar el estado de un empleado
@login_required  
def cambiar_estado(request, id):
    # Obtener el objeto Usuario completo
    # Import Estado locally to break potential circular dependencies
    from paginaPetrocentro.models import Estado

    usuario = get_object_or_404(Usuario, id=id)

    #----------------------------------------------------------------------------------------------------------------
    #obtener solo el estado del usuario
    id_estado1 = usuario.estado.id
    
    #comparar si el estado del usuario es uigual a dos
    #----------------------------------------------------------------------------------------------------------------
    if id_estado1 == 2:

      
        #generar una variable que es la que cambiara el estado 
        id_estado = 1
        
        #cambiar el estado de Id_estado con el registrado
        # Usamos get_or_create para asegurar que el estado exista. Si no, lo crea.
        estado, _ = Estado.objects.get_or_create(id=id_estado, defaults={'nombre': 'Activo'})
        
        #se agrega a el campo de estado en la base de datos 
        usuario.estado = estado
        
        #se guarda el estado del usuario en la base de datos
        usuario.save()
        
        if usuario:
            messages.success(request, f'Estado de {usuario.nombre} activado correctamente.')
        else:
            messages.error(request, 'No se ha cambiado el estado correctamente')
   
    #----------------------------------------------------------------------------------------------------------------
    else:
        user = usuario.user_id # Accedemos directamente al objeto User relacionado

# Cierra la sesión de usuarios cuya cuenta está deshabilitada
        if user.is_active:
            # Buscar las sesiones activas del usuario
            active_sessions =  Session.objects.filter(expire_date__gte=timezone.now())


            for session in active_sessions:
              # Si hay más de una sesión activa, se desloguea la última
                session_data = session.get_decoded()
                if session_data.get('_auth_user_id') == str(user.id):
                    messages.error(request,'cuenta deshabilitada')
                    # Eliminar la sesión activa
                    session.delete()
                 
             
              
        #generar una variable que es la que cambiara el estado        
        id_estado = 2
        
        #cambiar el estado de Id_estado con el registrado      
        # Usamos get_or_create para asegurar que el estado exista. Si no, lo crea.
        estado, _ = Estado.objects.get_or_create(id=id_estado, defaults={'nombre': 'Inactivo'})
        
        #se agrega a el campo de estado en la base de datos 
        usuario.estado = estado      
        #se guarda el estado del usuario en la base de datos 
        usuario.save()
        if usuario:
            messages.success(request, f'Estado de {usuario.nombre} desactivado correctamente.')
        else:
            messages.error(request, 'No se ha cambiado el estado correctamente')

    return redirect('empleados')

@login_required
def registrar_usuario(request):
    ctx = obtener_contexto_usuario(request)
    if not ctx['usuario']:
        return redirect('index')
    
    permisos = ctx['permisos']

    # Ahora, se verifica si el usuario tiene el permiso 'usuarios'
    # Import Estado locally to break potential circular dependencies
    from paginaPetrocentro.models import Estado

    if permisos.get('usuarios') == 1:
        # Lógica de la vista (GET)
        user_list = Usuario.objects.select_related('user_id', 'estado', 'empleado', 'empleado__id_cargo', 'empleado__id_rol').all().order_by('-user_id__date_joined')
        
        # --- FILTROS AVANZADOS ---
        busqueda = request.GET.get('buscar')
        filtro_estado = request.GET.get('estado')
        filtro_cargo = request.GET.get('cargo')
        filtro_fecha = request.GET.get('fecha')

        if busqueda:
            user_list = Usuario.objects.filter(
                Q(nombre__icontains=busqueda) |
                Q(correo__icontains=busqueda) |
                Q(empleado__id_cargo__nombre__icontains=busqueda)
            ).distinct()

        if filtro_estado and filtro_estado != '0':
            user_list = user_list.filter(estado_id=filtro_estado)
        
        if filtro_cargo and filtro_cargo != '0':
            user_list = user_list.filter(empleado__id_cargo_id=filtro_cargo)
            
        if filtro_fecha:
            user_list = user_list.filter(user_id__date_joined__date=filtro_fecha)

        form = RegisterForm(request.POST or None)
        p = Paginator(user_list, 10) # 10 por página
        page_number = request.GET.get('page')
        pagina = p.get_page(page_number)

        # Lógica de la vista (POST)
        if request.method == 'POST' and form.is_valid():
            username = form.cleaned_data.get('username')
            # ... validaciones de contraseña existentes en el form o aquí ...
            if form.cleaned_data.get('contraseña') != form.cleaned_data.get('confirmar_contraseña'):
                 messages.error(request, 'Las contraseñas no coinciden.')
            elif User.objects.filter(username=username).exists():
                 messages.error(request, 'El usuario ya existe.')
            else:
                 new_user = form.save()
                 usuario_creado = Usuario.objects.create(
                     user_id=new_user,
                     estado=Estado.objects.get(id=1),
                     nombre=form.cleaned_data['Nombre_completo'],
                     correo=form.cleaned_data['correo_electronico'],
                 )
                 
                 # --- CREACIÓN AUTOMÁTICA DE PERFIL DE EMPLEADO ---
                 rol_empleado, _ = Rol.objects.get_or_create(nombre="Empleado")
                 Empleado.objects.create(
                     id=usuario_creado.id,
                     user_id=new_user,
                     identificacion=0, # Valor temporal
                     telefono="0",      # Valor temporal
                     id_rol=rol_empleado,
                     fecha_ingreso=timezone.now().date(),
                     id_cargo=Cargo.objects.first(),
                     id_ubicacion=Ubicacion.objects.first(),
                     estado=usuario_creado.estado,
                     nombre=usuario_creado.nombre,
                     correo=usuario_creado.correo
                 )
                 
                 # Log de auditoría
                 try:
                     admin_resp = ctx['usuario']
                     LogActividad.objects.create(
                         usuario_afectado=usuario_creado,
                         admin_responsable=admin_resp,
                         accion="Creación de Usuario",
                         detalles=f"Usuario registrado manualmente por {admin_resp.nombre}"
                     )
                 except:
                     pass
                     
                 messages.success(request, 'Registro Exitoso.')
                 return redirect('usuarios')

        data = {
            **ctx,
            'form': form,
            'paginas': pagina,
            'paginator': p,
            # Datos para filtros y modales
            'cargos': Cargo.objects.all(),
            'roles': Rol.objects.all(),
            'filtros': {
                'estado': filtro_estado,
                'cargo': filtro_cargo,
                'fecha': filtro_fecha,
                'buscar': busqueda
            }
        }
        return render(request, 'dash/usuarios.html', data)
    
    # Caso 3: El usuario no tiene el permiso 'usuarios'.
    else:
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('index')

@login_required
def cambiar_estado_usuario(request, id):
    # Versión específica para la vista de usuarios
    if not request.user.is_superuser:
        messages.error(request, "Acción no autorizada.")
        return redirect('usuarios')

    usuario = get_object_or_404(Usuario, id=id)
    # Import Estado locally to break potential circular dependencies
    from paginaPetrocentro.models import Estado

    nuevo_estado_id = 2 if usuario.estado.id == 1 else 1
    nuevo_estado_nombre = "Inactivo" if nuevo_estado_id == 2 else "Activo"
    
    estado_obj, _ = Estado.objects.get_or_create(id=nuevo_estado_id, defaults={'nombre': nuevo_estado_nombre})
    usuario.estado = estado_obj
    usuario.save()

    # Si se desactiva, cerrar sesiones activas (User de Django)
    if nuevo_estado_id == 2:
        usuario.user_id.is_active = False
    else:
        usuario.user_id.is_active = True
    usuario.user_id.save()

    # Log
    try: # Use ctx['usuario'] for admin_resp
        admin_resp = obtener_contexto_usuario(request)['usuario']
        LogActividad.objects.create(
            usuario_afectado=usuario,
            admin_responsable=admin_resp,
            accion="Cambio de Estado",
            detalles=f"Estado cambiado a {nuevo_estado_nombre}"
        )
    except:
        pass

    messages.success(request, f"Estado de {usuario.nombre} actualizado a {nuevo_estado_nombre}.")
    return redirect('usuarios')

@login_required
def editar_usuario_rapido(request, id):
    ctx = obtener_contexto_usuario(request)
    if request.method == 'POST':
        usuario = get_object_or_404(Usuario, id=id)
        rol_id = request.POST.get('rol')
        cargo_id = request.POST.get('cargo')

        # Gestionar perfil de Empleado
        empleado, created = Empleado.objects.get_or_create(id=usuario.id, defaults={
            'user_id': usuario.user_id,
            'identificacion': 0, 'telefono': '0',
            'fecha_ingreso': timezone.now().date(),
            'id_ubicacion': Ubicacion.objects.first(), # Default fallback
            'id_cargo': Cargo.objects.first(), # Default fallback
            'estado': usuario.estado,
            'nombre': usuario.nombre, 'correo': usuario.correo
        })

        if rol_id:
            empleado.id_rol_id = rol_id
        if cargo_id:
            empleado.id_cargo_id = cargo_id
        
        empleado.save()
        
        # Log
        try: # Use ctx['usuario'] for admin_resp
            admin_resp = ctx['usuario']
            LogActividad.objects.create(
                usuario_afectado=usuario,
                admin_responsable=admin_resp,
                accion="Edición Rápida",
                detalles=f"Actualización de Rol/Cargo"
            )
        except:
            pass

        messages.success(request, "Usuario actualizado correctamente.")
    return redirect('usuarios')

@login_required
def reset_password_usuario(request, id):
    ctx = obtener_contexto_usuario(request)
    if request.method == 'POST':
        usuario = get_object_or_404(Usuario, id=id)
        new_pass = request.POST.get('new_password')
        
        if new_pass:
            user_django = usuario.user_id
            user_django.set_password(new_pass)
            user_django.save()
            
            # Log # Use ctx['usuario'] for admin_resp
            try:
                admin_resp = Usuario.objects.get(id=request.session.get('usuario_logeado'))
                LogActividad.objects.create(
                    usuario_afectado=usuario,
                    admin_responsable=admin_resp,
                    accion="Reset Contraseña",
                    detalles="Contraseña restablecida desde panel admin"
                )
            except:
                pass

            messages.success(request, f"Contraseña actualizada para {usuario.nombre}.")
    return redirect('usuarios')

@login_required
def exportar_usuarios_excel(request):
    # Reutilizar lógica de filtros del listado principal
    busqueda = request.GET.get('buscar')
    filtro_estado = request.GET.get('estado')
    filtro_cargo = request.GET.get('cargo')
    filtro_fecha = request.GET.get('fecha')

    user_list = Usuario.objects.select_related('empleado', 'empleado__id_cargo', 'empleado__id_rol').all()

    if busqueda:
        user_list = user_list.filter(Q(nombre__icontains=busqueda) | Q(correo__icontains=busqueda)).distinct()
    if filtro_estado and filtro_estado != '0':
        user_list = user_list.filter(estado_id=filtro_estado)
    if filtro_cargo and filtro_cargo != '0':
        user_list = user_list.filter(empleado__id_cargo_id=filtro_cargo)
    if filtro_fecha:
        user_list = user_list.filter(user_id__date_joined__date=filtro_fecha)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Reporte Usuarios'

    headers = ['ID', 'Nombre', 'Correo', 'Estado', 'Cargo', 'Rol', 'Fecha Registro']
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for u in user_list:
        cargo = u.empleado.id_cargo.nombre if hasattr(u, 'empleado') and u.empleado and u.empleado.id_cargo else 'N/A'
        rol_nombre = u.empleado.id_rol.nombre if hasattr(u, 'empleado') and u.empleado and u.empleado.id_rol else 'N/A'
        estado = u.estado.nombre if u.estado else 'Desconocido'
        fecha = u.user_id.date_joined.strftime('%Y-%m-%d') if u.user_id.date_joined else 'N/A'

        rol = rol_nombre
        
        sheet.append([u.id, u.nombre, u.correo, estado, cargo, rol, fecha])

    for col in ['B', 'C', 'E', 'F']: sheet.column_dimensions[col].autosize = True

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="usuarios_filtrados.xlsx"'
    workbook.save(response)
    return response

# -------------------------------- GESTIÓN DE TAREAS (Lógica Cargo/Rol) --------------------------------

@login_required
def gestion_tareas(request):
    ctx = obtener_contexto_usuario(request)

    # Acceso: Superusuario o empleado con perfil
    if not ctx['usuario']:
        messages.error(request, 'No se encontró tu perfil de usuario.')
        return redirect('index')
    
    # Si no es superusuario y no tiene perfil de empleado, no puede acceder
    if not request.user.is_superuser and not ctx['empleado']:
            messages.error(request, 'No tienes un perfil de empleado para acceder a esta página.')
            return redirect('index')

    # 2. Filtrar Tareas según CARGO (Tareas Concretas)
    if ctx['nombre_rol'] == "ADMINISTRADOR" or request.user.is_superuser:
        # El administrador (Rol) tiene acceso global, ve todas las tareas de todos los cargos
        tareas = Tarea.objects.select_related('empleado', 'id_cargo').all()
        
        # Filtros para el admin
        empleado_filtro = request.GET.get('empleado_filtro')
        estado_filtro = request.GET.get('estado_tarea')

        if empleado_filtro:
            tareas = tareas.filter(empleado__id=empleado_filtro)
        
        if estado_filtro == 'completada':
            tareas = tareas.filter(completada=True)
        elif estado_filtro == 'pendiente':
            tareas = tareas.filter(completada=False)

        tareas = tareas.order_by('-fecha_creacion')
        lista_completa_empleados = Empleado.objects.all().order_by('nombre')
    else:
        # El empleado ve SOLO las tareas asignadas a su EMPLEADO específico
        if ctx['empleado']:
            tareas = Tarea.objects.filter(empleado=ctx['empleado']).order_by('fecha_limite')
        else:
            tareas = Tarea.objects.none()
        lista_completa_empleados = Empleado.objects.none()

    # Determinar si el usuario tiene permiso para crear tareas
    is_manager_cargo = False
    if ctx['empleado'] and ctx['empleado'].id_cargo:
        is_manager_cargo = ctx['empleado'].id_cargo.nombre in [
            "Gerente de Operaciones", 
            "Administrador de Contrato", 
            "Coordinador de Operaciones",
            "Coordinador HSEQ"
        ]
    can_create_tasks = ctx.get('crear') and (request.user.is_superuser or is_manager_cargo)

    # 3. Lógica para Crear Tarea (Solo si tiene permiso por Rol y Cargo) # Use ctx['crear']
    if request.method == 'POST':
        if can_create_tasks:
            titulo = request.POST.get('titulo')
            descripcion = request.POST.get('descripcion')
            fecha_limite = request.POST.get('fecha_limite')
            id_cargo_destino = request.POST.get('id_cargo')
            id_empleado_destino = request.POST.get('id_empleado')
            
            if id_empleado_destino:
                # Opción A: Asignar a un empleado específico
                try:
                    emp = Empleado.objects.get(id=id_empleado_destino)
                    Tarea.objects.create(
                        titulo=titulo,
                        descripcion=descripcion,
                        fecha_limite=fecha_limite,
                        id_cargo=emp.id_cargo, 
                        empleado=emp
                    )
                    # Notificación por correo
                    try:
                        send_mail(
                            'Nueva Tarea Asignada - Petrocentro',
                            f'Hola {emp.nombre},\n\nSe te ha asignado una nueva tarea: "{titulo}".\nFecha límite: {fecha_limite}\n\nDescripción: {descripcion}\n\nPor favor ingresa a tu agenda para gestionarla.',
                            settings.EMAIL_HOST_USER,
                            [emp.correo],
                            fail_silently=True
                        )
                    except:
                        pass
                    messages.success(request, f'Tarea asignada correctamente a {emp.nombre}.')
                except Empleado.DoesNotExist:
                    messages.error(request, 'El empleado seleccionado no existe.')
                    
            elif id_cargo_destino:
                # Opción B: Asignar masivamente por cargo
                cargo = Cargo.objects.get(id_cargo=id_cargo_destino)
                empleados = Empleado.objects.filter(id_cargo=cargo)
                
                # Creamos una tarea individual para CADA empleado del cargo
                count = 0
                for emp in empleados:
                    Tarea.objects.create(
                        titulo=titulo,
                        descripcion=descripcion,
                        fecha_limite=fecha_limite,
                        id_cargo=cargo,
                        empleado=emp # Asignación individual
                    )
                    # Notificación por correo (individual en bucle)
                    try:
                        send_mail(
                            'Nueva Tarea Asignada - Petrocentro',
                            f'Hola {emp.nombre},\n\nSe te ha asignado una nueva tarea: "{titulo}".\nFecha límite: {fecha_limite}\n\nDescripción: {descripcion}\n\nPor favor ingresa a tu agenda para gestionarla.',
                            settings.EMAIL_HOST_USER,
                            [emp.correo],
                            fail_silently=True
                        )
                    except:
                        pass
                    count += 1
                
                if count > 0:
                    messages.success(request, f'Se ha asignado la tarea a {count} empleados del cargo {cargo.nombre}.')
                else:
                    messages.warning(request, f'El cargo {cargo.nombre} no tiene empleados activos actualmente.')
            
            else:
                messages.error(request, 'Debes seleccionar un Cargo o un Empleado para asignar la tarea.')
                
        else:
            messages.error(request, 'No tienes los permisos o el cargo adecuado para crear tareas.')
        
        return redirect('gestion_tareas')

    data = {
        **ctx, # Unpack the common context
        'tareas': tareas,
        'cargos': Cargo.objects.all(), # Para el select de crear tarea
        'empleados_filtro': lista_completa_empleados,
        'can_create_tasks': can_create_tasks,
    }
    return render(request, 'plantillas/tareas.html', data)

@login_required
def completar_tarea(request, id_tarea):
    if request.method == 'POST':
        tarea = get_object_or_404(Tarea, id_tarea=id_tarea)
        # Security check: only owner or admin can complete
        usuario_logeado_id = request.session.get('usuario_logeado')
        if not request.user.is_superuser and tarea.empleado and tarea.empleado.id != usuario_logeado_id: # Added check for tarea.empleado
            messages.error(request, "No tienes permiso para modificar esta tarea o no está asignada a ti.")
            return redirect('agenda_personal')

        archivo = request.FILES.get('archivo_adjunto')
        comentarios = request.POST.get('comentarios_cumplimiento')

        if archivo:
            # Si ya existe un archivo, lo borramos antes de guardar el nuevo.
            if tarea.archivo_adjunto:
                tarea.archivo_adjunto.delete(save=False)
            tarea.archivo_adjunto = archivo
        
        if comentarios:
            tarea.comentarios_cumplimiento = comentarios

        tarea.completada = True
        tarea.save()
        messages.success(request, f"Tarea '{tarea.titulo}' completada exitosamente.")
    
    # Redirect logic: if admin, go to task management, if employee, go to personal agenda.
    referer = request.META.get('HTTP_REFERER')
    if request.user.is_superuser or (referer and 'ver_agenda' in referer):
        return redirect(referer or 'gestion_tareas')
    else:
        return redirect('agenda_personal')

@login_required
def reactivar_tarea(request, id_tarea):
    ctx = obtener_contexto_usuario(request)
    if not request.user.is_superuser and not ctx['editar']: # Check for edit permission
        messages.error(request, "No tienes permisos para reactivar tareas.")
        return redirect('agenda_personal')

    tarea = get_object_or_404(Tarea, id_tarea=id_tarea)
    tarea.completada = False
    if tarea.archivo_adjunto:
        tarea.archivo_adjunto.delete(save=False) # Borra el archivo del almacenamiento
    tarea.save()
    messages.info(request, f"Tarea '{tarea.titulo}' ha sido reactivada.")
    return redirect('gestion_tareas')

@login_required
def eliminar_tarea(request, id_tarea):
    ctx = obtener_contexto_usuario(request)

    if ctx['eliminar']:
        tarea = get_object_or_404(Tarea, id_tarea=id_tarea)
        tarea.delete()
        messages.success(request, 'La tarea ha sido eliminada correctamente.')
    else:
        messages.error(request, 'No tienes permisos para eliminar tareas.')
    
    return redirect('gestion_tareas')

# -------------------------------- GESTIÓN DE CURSOS Y AGENDA PERSONAL --------------------------------

@login_required
def agenda_personal(request):
    ctx = obtener_contexto_usuario(request)
    
    # 1. Obtener el perfil del empleado
    # Si es superusuario pero no empleado, redirigir a una vista de admin
    if not ctx['empleado']:
        if request.user.is_superuser:
            return redirect('gestion_cursos')
        messages.error(request, 'No se encontró tu perfil de empleado.')
        return redirect('index')

    # 2. Lógica de asignación automática de cursos obligatorios
    if ctx['empleado'].id_cargo:
        cursos_obligatorios = Curso.objects.filter(obligatorio_para=ctx['empleado'].id_cargo)
        
        # Optimización: Obtener IDs existentes primero para evitar consultas múltiples
        cursos_asignados_ids = EmpleadoCurso.objects.filter(
            empleado=ctx['empleado'],
            curso__in=cursos_obligatorios
        ).values_list('curso_id', flat=True)
        
        nuevos_cursos = []
        for curso in cursos_obligatorios:
            if curso.id_curso not in cursos_asignados_ids:
                nuevos_cursos.append(EmpleadoCurso(empleado=ctx['empleado'], curso=curso))
        
        if nuevos_cursos:
            EmpleadoCurso.objects.bulk_create(nuevos_cursos)

    # 3. Obtener Tareas y Cursos para la agenda
    # Filtramos tareas asignadas ESPECÍFICAMENTE a este empleado
    tareas_qs = Tarea.objects.filter(empleado=ctx['empleado']).order_by('fecha_limite')
    
    # --- Separación Pendientes vs Historial ---
    tareas_pendientes = [t for t in tareas_qs if not t.completada]
    tareas_historial = [t for t in tareas_qs if t.completada]
    
    # --- NOTIFICACIONES AUTOMÁTICAS ---
    today = timezone.now().date()
    
    # Alerta si hay tareas vencidas
    tareas_vencidas = sum(1 for t in tareas_pendientes if t.fecha_limite < today)
    if tareas_vencidas > 0:
        messages.warning(request, f'¡Atención! Tienes {tareas_vencidas} tareas operativas vencidas. Por favor gestiónalas.')
        
    # Alerta de tareas próximas a vencer (Próximos 3 días)
    limit_date = today + timedelta(days=3)
    tareas_proximas = sum(1 for t in tareas_pendientes if today <= t.fecha_limite <= limit_date)
    if tareas_proximas > 0:
        messages.info(request, f'Tienes {tareas_proximas} tareas que vencen en los próximos 3 días.')

    cursos_qs = EmpleadoCurso.objects.filter(empleado=ctx['empleado']).select_related('curso').order_by('curso__nombre')
    
    cursos_pendientes = [c for c in cursos_qs if c.estado != 'APROBADO']
    cursos_historial = [c for c in cursos_qs if c.estado == 'APROBADO']

    # Solicitudes de inventario del supervisor (para que aparezcan en su perfil/agenda)
    mis_solicitudes = TransferRequest.objects.filter(supervisor=ctx['usuario']).exclude(status='PENDING').select_related('processed_by').order_by('-processed_at')[:10]

    # --- ALERTAS DE CURSOS Y CERTIFICADOS ---
    # Alerta de cursos vencidos o próximos a vencer (en Historial/Aprobados)
    certs_por_vencer = 0
    for c in cursos_historial:
        if c.fecha_vencimiento and today <= c.fecha_vencimiento <= limit_date:
            certs_por_vencer += 1
    
    if certs_por_vencer > 0:
        messages.warning(request, f'¡Aviso! Tienes {certs_por_vencer} certificado(s) próximo(s) a vencer o vencidos.')
        
    # Alerta específica para cursos RECHAZADOS (Feedback de estado)
    rechazados_count = sum(1 for c in cursos_pendientes if c.estado == 'RECHAZADO')
    if rechazados_count > 0:
        messages.error(request, f'Atención: Tienes {rechazados_count} certificado(s) rechazado(s). Por favor revisa las observaciones y vuelve a subirlo.')

    # Alerta genérica de cursos pendientes (para motivar el avance)
    num_pendientes = len(cursos_pendientes)
    if num_pendientes > 0:
        messages.info(request, f'Recordatorio: Tienes {num_pendientes} cursos obligatorios pendientes por aprobar.')

    # --- LOGICA DE GAMIFICACIÓN ---
    total_asignados = len(cursos_qs)
    progreso_cumplimiento = 0
    if total_asignados > 0:
        progreso_cumplimiento = int((len(cursos_historial) / total_asignados) * 100)
    
    # --- CÁLCULO CUMPLIMIENTO TAREAS ---
    total_tareas = len(tareas_qs)
    progreso_tareas = 0
    if total_tareas > 0:
        progreso_tareas = int((len(tareas_historial) / total_tareas) * 100)
    
    # --- GAMIFICACIÓN AVANZADA ---
    # 1. Ranking por Área (Ubicación)
    ranking = 0
    total_area = 0
    if ctx['empleado'].id_ubicacion:
        # Optimizamos el cálculo del ranking usando anotaciones para evitar el problema N+1
        coleagas = Empleado.objects.filter(id_ubicacion=ctx['empleado'].id_ubicacion).annotate(
            c_ok=Count('cursos_asignados', filter=Q(cursos_asignados__estado='APROBADO'), distinct=True),
            t_ok=Count('tarea', filter=Q(tarea__completada=True), distinct=True)
        )
        total_area = coleagas.count()
        # Calculamos puntaje simple: (Cursos Aprobados * 10) + (Tareas Completadas * 5)
        scores = []
        for col in coleagas:
            puntos = (col.c_ok * 10) + (col.t_ok * 5)
            scores.append(puntos)
        
        scores.sort(reverse=True) # Ordenar descendente
        # Mi puntaje
        mi_puntaje = (len(cursos_historial) * 10) + (len(tareas_historial) * 5)
        # Encontrar mi posición (índice + 1)
        # Manejo simple de empates: tomamos la primera aparición
        ranking = scores.index(mi_puntaje) + 1 if mi_puntaje in scores else total_area

    # --- RETO MENSUAL (GAMIFICACIÓN) ---
    # Reto: "Operación Impecable" - Sin tareas vencidas y al menos 1 completada
    reto_mensual = {
        'titulo': 'Operación Impecable',
        'descripcion': 'Completa tareas este mes sin tener ninguna vencida.',
        'cumplido': False,
        'icono': 'fa-solid fa-shield-halved'
    }
    tareas_completadas_mes = [t for t in tareas_historial if t.fecha_creacion.month == today.month]
    tiene_vencidas = any(t.is_overdue for t in tareas_pendientes)
    if not tiene_vencidas and len(tareas_completadas_mes) > 0:
        reto_mensual['cumplido'] = True

    insignias = []
    # Insignia por Cumplimiento de Cursos
    if progreso_cumplimiento == 100 and total_asignados > 0:
        insignias.append({'icono': 'fa-solid fa-medal', 'color': 'text-warning', 'titulo': 'Excelencia Académica', 'desc': '100% de cursos aprobados'})
    elif progreso_cumplimiento >= 50:
        insignias.append({'icono': 'fa-solid fa-star-half-stroke', 'color': 'text-info', 'titulo': 'Buen Avance', 'desc': 'Más del 50% completado'})
    
    # Insignia por Tareas al día
    if not tareas_pendientes and tareas_historial:
        insignias.append({'icono': 'fa-solid fa-clipboard-check', 'color': 'text-success', 'titulo': 'Operativo Eficaz', 'desc': 'Sin tareas pendientes'})
    
    # Insignia de Bienvenida (si acaba de empezar)
    if progreso_cumplimiento == 0 and not tareas_historial:
        insignias.append({'icono': 'fa-regular fa-hand-peace', 'color': 'text-primary', 'titulo': 'Nuevo Ingreso', 'desc': '¡Bienvenido al equipo!'})
    # ------------------------------

    context = {
        **ctx, # Unpack the common context
        'tareas_pendientes': tareas_pendientes,
        'tareas_historial': tareas_historial,
        'cursos_pendientes': cursos_pendientes,
        'cursos_historial': cursos_historial,
        'tareas': tareas_qs, # Necesario para generar los modales de todas las tareas
        # 'usuario': empleado_profile, # Empleado hereda de Usuario, already in ctx
        # 'empleado': empleado_profile, # already in ctx
        # 'nombre_rol': nombre_rol, # already in ctx
        # 'permisos': permisos, # already in ctx
        # 'inventario': permisos.get('inventario', 0), # already in ctx
        # 'usuarios': permisos.get('usuarios', 0), # already in ctx
        'today': today, # Añadimos la fecha de hoy para comparaciones en la plantilla
        'progreso_cumplimiento': progreso_cumplimiento,
        'insignias': insignias,
        'total_cursos': total_asignados,
        'progreso_tareas': progreso_tareas,
        'total_tareas': total_tareas,
        'mis_solicitudes': mis_solicitudes,
        # Datos Gamificación
        'ranking_area': ranking,
        'total_area': total_area,
        'reto_mensual': reto_mensual,
    }
    
    return render(request, 'plantillas/agenda.html', context)

@login_required
def subir_certificado(request, id_empleado_curso):
    ctx = obtener_contexto_usuario(request)
    if request.method == 'POST':
        empleado_curso = get_object_or_404(EmpleadoCurso, id_empleado_curso=id_empleado_curso)
        
        # Verificar que el usuario que sube es el dueño del curso
        if empleado_curso.empleado.id != ctx['usuario'].id:
            messages.error(request, "No tienes permiso para modificar este curso.")
            return redirect('agenda_personal')

        certificado_file = request.FILES.get('certificado')
        if certificado_file:
            # Borra el certificado anterior si existe
            if empleado_curso.certificado:
                empleado_curso.certificado.delete(save=False)

            empleado_curso.certificado = certificado_file
            empleado_curso.estado = 'EN_REVISION' # Cambia a estado de revisión
            empleado_curso.save()
            
            # Notificar al Administrador que llegó un nuevo certificado
            try:
                send_mail(
                    f"Certificado para Validar: {empleado_curso.empleado.nombre}",
                    f"El empleado {empleado_curso.empleado.nombre} ha enviado su certificado del curso '{empleado_curso.curso.nombre}'.\n\nPor favor ingrese al panel de 'Gestión de Cursos' para aprobarlo o rechazarlo.",
                    settings.EMAIL_HOST_USER,
                    [settings.EMAIL_HOST_USER], # Se envía al correo principal del sistema (Admin)
                    fail_silently=True
                )
            except: pass

            messages.success(request, f"Tu certificado para '{empleado_curso.curso.nombre}' fue enviado para validación.")
        else:
            messages.error(request, "Debes seleccionar un archivo.")

    return redirect('agenda_personal')

@login_required
def eliminar_certificado(request, id_empleado_curso):
    ctx = obtener_contexto_usuario(request)
    if request.method == 'POST':
        empleado_curso = get_object_or_404(EmpleadoCurso, id_empleado_curso=id_empleado_curso)
        
        if empleado_curso.empleado.id != ctx['usuario'].id:
            messages.error(request, "No tienes permiso para modificar este curso.")
            return redirect('agenda_personal')

        if empleado_curso.certificado:
            empleado_curso.certificado.delete(save=False)
            empleado_curso.certificado = None
            empleado_curso.estado = 'PENDIENTE' # Reiniciamos el estado para que pueda subir uno nuevo
            empleado_curso.save()
            messages.success(request, f"Archivo eliminado. El curso vuelve a estado Pendiente.")
    
    return redirect('agenda_personal')

@login_required
def soporte_empleado(request):
    """Procesa el formulario de soporte enviado por el empleado."""
    ctx = obtener_contexto_usuario(request)
    if request.method == 'POST':
        asunto = request.POST.get('asunto')
        mensaje = request.POST.get('mensaje')
        usuario = ctx['usuario'] # Usuario de Django (auth_user)
        
        # Preparamos el cuerpo del correo
        cuerpo_correo = f"""
        Solicitud de Soporte enviada desde el Portal de Empleados.
        
        Usuario: {usuario.nombre}
        Correo: {usuario.correo}
        
        Mensaje:
        {mensaje}
        """
        
        try:
            # Enviar correo al administrador (usando el correo configurado en settings)
            send_mail(
                subject=f"Soporte Empleado: {asunto}",
                message=cuerpo_correo,
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[settings.EMAIL_HOST_USER], # Se envía al admin
                fail_silently=False,
            )
            messages.success(request, 'Tu solicitud de soporte ha sido enviada al administrador.')
        except Exception as e:
            print(f"Error enviando soporte: {e}")
            messages.error(request, 'Hubo un error al enviar tu solicitud. Por favor intenta más tarde o contacta por otro medio.')
            
    return redirect('agenda_personal')

@login_required
def gestion_cursos(request):
    ctx = obtener_contexto_usuario(request)

    # Verificación de permisos de Administrador
    # Acceso permitido a superusuarios o a empleados con permiso 'consultar' en inventario o usuarios
    if not request.user.is_superuser and not (ctx['consultar'] or ctx['inventario'] or ctx['usuarios']):
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('index')

    # Lógica para crear/editar un curso
    if request.method == 'POST':
        # Validación de seguridad extra para POST
        if not (ctx['crear'] or ctx['editar']):
            messages.error(request, 'No tienes permisos para realizar cambios.')
            return redirect('gestion_cursos')

        nombre_curso = request.POST.get('nombre_curso')
        descripcion_curso = request.POST.get('descripcion_curso')
        external_url = request.POST.get('external_url')
        modalidad = request.POST.get('modalidad')
        lugar = request.POST.get('lugar')
        periodicidad = request.POST.get('periodicidad')
        cargos_obligatorios_ids = request.POST.getlist('cargos_obligatorios')

        if nombre_curso:
            curso, created = Curso.objects.get_or_create(nombre=nombre_curso)
            curso.descripcion = descripcion_curso
            if modalidad:
                curso.modalidad = modalidad
            curso.lugar = lugar # Guardar lugar (puede ser vacío si es virtual)
            if periodicidad:
                curso.periodicidad = periodicidad
            curso.external_url = external_url
            curso.obligatorio_para.set(cargos_obligatorios_ids)
            curso.save()
            
            # Asignar automáticamente el curso a los empleados existentes de esos cargos
            empleados_afectados = Empleado.objects.filter(id_cargo__in=cargos_obligatorios_ids)
            for emp in empleados_afectados:
                obj, created = EmpleadoCurso.objects.get_or_create(empleado=emp, curso=curso)
                if created:
                    # Notificar nueva asignación
                    try:
                        send_mail(
                            'Nuevo Curso Asignado - Petrocentro',
                            f'Hola {emp.nombre},\n\nSe te ha asignado el curso obligatorio "{nombre_curso}". Por favor ingresa a tu agenda para gestionarlo.\n\nSaludos.',
                            settings.EMAIL_HOST_USER,
                            [emp.correo],
                            fail_silently=True
                        )
                    except:
                        pass

            messages.success(request, f"Curso '{nombre_curso}' guardado y asignado correctamente.")
            return redirect('gestion_cursos')

    # Lógica para mostrar la vista
    cursos = Curso.objects.prefetch_related('obligatorio_para').all()
    cargos = Cargo.objects.all()
    certificados_pendientes = EmpleadoCurso.objects.filter(estado='EN_REVISION').select_related('empleado', 'curso', 'empleado__user_id')
    
    # --- RECORDATORIOS DE RENOVACIÓN (Cursos por Vencer o Vencidos) ---
    fecha_limite_recordatorio = timezone.now().date() + timedelta(days=30) # Vencen en los próximos 30 días o ya vencieron
    renovaciones_pendientes = EmpleadoCurso.objects.filter(estado='APROBADO', fecha_vencimiento__lte=fecha_limite_recordatorio).select_related('empleado', 'curso', 'empleado__id_cargo').order_by('fecha_vencimiento')

    
    # --- Filtros Avanzados para Certificados Pendientes ---
    pend_curso = request.GET.get('pend_curso')
    pend_cargo = request.GET.get('pend_cargo')
    
    if pend_curso:
        certificados_pendientes = certificados_pendientes.filter(curso__id_curso=pend_curso)
    if pend_cargo:
        certificados_pendientes = certificados_pendientes.filter(empleado__id_cargo__id_cargo=pend_cargo)

    # --- INDICADORES DASHBOARD (KPIs) ---
    # Se calculan totales globales para gráficas y métricas
    total_asignaciones = EmpleadoCurso.objects.count()
    total_aprobados = EmpleadoCurso.objects.filter(estado='APROBADO').count()
    porcentaje_cumplimiento = (total_aprobados / total_asignaciones * 100) if total_asignaciones > 0 else 0
    pendientes_revision_count = EmpleadoCurso.objects.filter(estado='EN_REVISION').count()
    cursos_activos_count = Curso.objects.count()
    pendientes_realizar = EmpleadoCurso.objects.filter(estado='PENDIENTE').count()

    # --- INDICADORES TAREAS (NUEVO PARA DASHBOARD GLOBAL) ---
    total_tareas_asignadas = Tarea.objects.count()
    tareas_completadas_count = Tarea.objects.filter(completada=True).count()
    tareas_pendientes_count = Tarea.objects.filter(completada=False).count()
    cumplimiento_tareas_global = (tareas_completadas_count / total_tareas_asignadas * 100) if total_tareas_asignadas > 0 else 0

    # --- DATOS PARA GRÁFICAS (Chart.js) ---
    # 1. Estado Global (Pie Chart)
    estados_agrupados = EmpleadoCurso.objects.values('estado').annotate(total=Count('estado'))
    datos_estados = {est['estado']: est['total'] for est in estados_agrupados}
    
    grafica_global = {
        'labels': ['Aprobado (Cumplido)', 'En Revisión', 'Pendiente', 'Rechazado'],
        'data': [
            datos_estados.get('APROBADO', 0),
            datos_estados.get('EN_REVISION', 0),
            datos_estados.get('PENDIENTE', 0),
            datos_estados.get('RECHAZADO', 0)
        ],
    }

    # 2. Cumplimiento por Cargo (Bar Chart)
    cumplimiento_cargos = EmpleadoCurso.objects.values('empleado__id_cargo__nombre').annotate(
        total=Count('id_empleado_curso'),
        aprobados=Count(Case(When(estado='APROBADO', then=1), output_field=IntegerField()))
    ).order_by('empleado__id_cargo__nombre').exclude(empleado__id_cargo__nombre__isnull=True)

    cargos_labels = []
    cargos_data = []
    for c in cumplimiento_cargos:
        cargos_labels.append(c['empleado__id_cargo__nombre'])
        val = round((c['aprobados'] / c['total'] * 100), 1) if c['total'] > 0 else 0
        cargos_data.append(val)

    # 3. Cumplimiento por Área (Ubicación)
    cumplimiento_areas = EmpleadoCurso.objects.values('empleado__id_ubicacion__nombre').annotate(
        total=Count('id_empleado_curso'),
        aprobados=Count(Case(When(estado='APROBADO', then=1), output_field=IntegerField()))
    ).order_by('empleado__id_ubicacion__nombre').exclude(empleado__id_ubicacion__nombre__isnull=True)

    areas_labels = []
    areas_data = []
    for a in cumplimiento_areas:
        areas_labels.append(a['empleado__id_ubicacion__nombre'])
        val = round((a['aprobados'] / a['total'] * 100), 1) if a['total'] > 0 else 0
        areas_data.append(val)

    # --- Lógica de Supervisión (Matriz de Cumplimiento) ---
    # Filtros para la tabla general de seguimiento
    estado_filtro = request.GET.get('estado')
    curso_filtro = request.GET.get('curso')
    cargo_filtro = request.GET.get('cargo')
    
    seguimiento = EmpleadoCurso.objects.select_related('empleado', 'curso', 'empleado__id_cargo').all()
    
    if estado_filtro:
        seguimiento = seguimiento.filter(estado=estado_filtro)
    if curso_filtro:
        seguimiento = seguimiento.filter(curso__id_curso=curso_filtro)
    if cargo_filtro:
        seguimiento = seguimiento.filter(empleado__id_cargo_id=cargo_filtro)
        
    seguimiento = seguimiento.order_by('empleado__nombre', 'curso__nombre')

    context = {
        **ctx, # Unpack the common context
        'cursos': cursos,
        'cargos': cargos,
        'certificados_pendientes': certificados_pendientes,
        'renovaciones_pendientes': renovaciones_pendientes,
        'seguimiento': seguimiento,
        'stats': {
            'cumplimiento': round(porcentaje_cumplimiento, 1),
            'pendientes_revision': pendientes_revision_count,
            'total_asignaciones': total_asignaciones,
            'cursos_activos': cursos_activos_count,
            'pendientes_realizar': pendientes_realizar,
            # KPIs de Tareas
            'total_tareas': total_tareas_asignadas,
            'cumplimiento_tareas': round(cumplimiento_tareas_global, 1),
            'tareas_pendientes': tareas_pendientes_count
        },
        'graficas': {
            'global': json.dumps(grafica_global),
            'cargos_labels': json.dumps(cargos_labels),
            'cargos_data': json.dumps(cargos_data),
            # Datos para gráfica de áreas
            'areas_labels': json.dumps(areas_labels),
            'areas_data': json.dumps(areas_data),
        },
        'filtros_pendientes': {'curso': pend_curso, 'cargo': pend_cargo},
        # 'usuario': usuario_profile, # Already in ctx
        # 'empleado': empleado_profile, # Already in ctx
        # 'nombre_rol': "Administrador", # Already in ctx
        # 'permisos': permisos, # Already in ctx
        # 'usuarios': permisos.get('usuarios', 0), # Already in ctx
        # 'inventario': permisos.get('inventario', 0), # Already in ctx
    }
    return render(request, 'plantillas/gestion_cursos_obligatorios.html', context)

@login_required
def validar_certificado(request, id_empleado_curso):
    ctx = obtener_contexto_usuario(request)
    # --- Verificación de Permisos (Rol Granular) ---
    # Se permite validar a Superusuarios O a empleados con permiso 'Editar' (ej. Supervisor/Validador)
    can_validate = False
    
    if request.user.is_superuser or ctx['editar']:
        can_validate = True
            
    if not can_validate:
        messages.error(request, "No tienes permiso para validar certificados.")
        return redirect('gestion_cursos')

    if request.method == 'POST':
        empleado_curso = get_object_or_404(EmpleadoCurso, id_empleado_curso=id_empleado_curso)
        accion = request.POST.get('accion')
        comentarios = request.POST.get('comentarios', '')

        if accion == 'aprobar':
            empleado_curso.estado = 'APROBADO'
            empleado_curso.comentarios_revision = ''
            
            # Calcular fecha de vencimiento automática según periodicidad
            if empleado_curso.curso.periodicidad == 'ANUAL':
                empleado_curso.fecha_vencimiento = timezone.now().date() + timedelta(days=365)
            elif empleado_curso.curso.periodicidad == 'SEMESTRAL':
                empleado_curso.fecha_vencimiento = timezone.now().date() + timedelta(days=180)
            elif empleado_curso.curso.periodicidad == 'MENSUAL':
                empleado_curso.fecha_vencimiento = timezone.now().date() + timedelta(days=30)
            # Si es UNICO, se mantiene la fecha que tuviera o null
            
            empleado_curso.save()
            
            # Notificación Automática al Empleado
            try:
                send_mail(
                    'Certificado Aprobado - Petrocentro',
                    f'Hola {empleado_curso.empleado.nombre},\n\nTu certificado del curso "{empleado_curso.curso.nombre}" ha sido aprobado exitosamente.\n\nAtentamente,\nEquipo de Formación.',
                    settings.EMAIL_HOST_USER,
                    [empleado_curso.empleado.correo],
                    fail_silently=True
                )
            except Exception:
                pass
                
            messages.success(request, f"Certificado de {empleado_curso.empleado.nombre} APROBADO.")
        elif accion == 'rechazar':
            empleado_curso.estado = 'RECHAZADO'
            empleado_curso.comentarios_revision = comentarios
            if empleado_curso.certificado:
                empleado_curso.certificado.delete(save=False)
            empleado_curso.save()
            
            # Notificación Automática de Rechazo
            try:
                send_mail(
                    'Certificado Rechazado - Petrocentro',
                    f'Hola {empleado_curso.empleado.nombre},\n\nTu certificado del curso "{empleado_curso.curso.nombre}" ha sido rechazado.\n\nMotivo: {comentarios}\n\nPor favor ingresa a la plataforma y sube el documento correcto.',
                    settings.EMAIL_HOST_USER,
                    [empleado_curso.empleado.correo],
                    fail_silently=True
                )
            except Exception:
                pass
                
            messages.warning(request, f"Certificado de {empleado_curso.empleado.nombre} RECHAZADO.")
    
    return redirect('gestion_cursos')

@login_required
def editar_curso(request, id_curso):
    ctx = obtener_contexto_usuario(request)
    # Verificación de permisos básica (similar a gestion_cursos)
    if not request.user.is_superuser and not ctx['editar']:
        messages.error(request, 'No tienes permisos.')
        return redirect('gestion_cursos')

    curso = get_object_or_404(Curso, id_curso=id_curso)
    
    if request.method == 'POST':
        nombre_curso = request.POST.get('nombre_curso')
        descripcion_curso = request.POST.get('descripcion_curso')
        external_url = request.POST.get('external_url')
        modalidad = request.POST.get('modalidad')
        lugar = request.POST.get('lugar')
        periodicidad = request.POST.get('periodicidad')
        cargos_obligatorios_ids = request.POST.getlist('cargos_obligatorios')

        if nombre_curso:
            curso.nombre = nombre_curso
            curso.descripcion = descripcion_curso
            if modalidad:
                curso.modalidad = modalidad
            curso.lugar = lugar
            if periodicidad:
                curso.periodicidad = periodicidad
            curso.external_url = external_url
            curso.obligatorio_para.set(cargos_obligatorios_ids)
            curso.save()
            
            # Actualizar asignaciones a empleados
            empleados_afectados = Empleado.objects.filter(id_cargo__in=cargos_obligatorios_ids)
            emails_count = 0
            for emp in empleados_afectados:
                obj, created = EmpleadoCurso.objects.get_or_create(empleado=emp, curso=curso)
                if created:
                    # Notificar nueva asignación
                    try:
                        send_mail(
                            'Nuevo Curso Asignado - Petrocentro',
                            f'Hola {emp.nombre},\n\nSe te ha asignado el curso obligatorio "{curso.nombre}". Por favor ingresa a tu agenda para gestionarlo.\n\nSaludos.',
                            settings.EMAIL_HOST_USER,
                            [emp.correo],
                            fail_silently=True
                        )
                    except:
                        pass
            
            messages.success(request, f"Curso '{nombre_curso}' actualizado correctamente.")
            return redirect('gestion_cursos')

    # Contexto para el formulario de edición
    context = {
        **ctx, # Unpack the common context
        'curso': curso,
        'cargos': Cargo.objects.all(),
        # 'usuario': get_object_or_404(Usuario, id=request.session.get('usuario_logeado')), # Corregido para foto perfil, already in ctx
        # 'usuarios': 1, # Admin tiene permisos de usuario por defecto, already in ctx
        # 'nombre_rol': "Administrador", # already in ctx
        # Pre-seleccionar cargos
        'cargos_seleccionados': curso.obligatorio_para.values_list('id_cargo', flat=True)
    }
    return render(request, 'plantillas/editar_curso.html', context)

@login_required
def eliminar_curso(request, id_curso):
    ctx = obtener_contexto_usuario(request)
    # Verificación de permisos
    if not request.user.is_superuser and not ctx['eliminar']:
        messages.error(request, "No tienes permisos para eliminar cursos.")
        return redirect('gestion_cursos')

    curso = get_object_or_404(Curso, id_curso=id_curso)
    
    try:
        curso.delete()
        messages.success(request, 'Curso eliminado correctamente.')
    except Exception as e:
        messages.error(request, f'Error al eliminar el curso: {e}')
    
    return redirect('gestion_cursos')

# -------------------------------- PERFILES DE ROL (REDIRECCIÓN Y VISTAS PRINCIPALES) --------------------------------

@login_required
def perfil_empleado(request):
    """Vista principal para el rol de Empleado: Muestra Agenda (Tareas + Cursos)"""
    return agenda_personal(request)

@login_required
def perfil_admin(request):
    """Vista principal para el rol de Administrador: Herramientas de Gestión (Cursos, Certificados)"""
    return gestion_cursos(request)

@login_required
def ver_agenda_empleado(request, empleado_id):
    """
    Vista para que un Administrador vea la agenda de un empleado específico.
    """
    ctx = obtener_contexto_usuario(request)

    # 1. Check if logged-in user is an admin
    if not request.user.is_superuser:
        messages.error(request, "No tienes permisos para ver la agenda de otros empleados.")
        return redirect('empleados')

    # 2. Get the target employee's profile
    try:
        empleado_profile = Empleado.objects.select_related('id_cargo', 'id_rol', 'user_id').get(id=empleado_id)
    except Empleado.DoesNotExist:
        messages.error(request, 'El empleado solicitado no existe.')
        return redirect('empleados')

    # 3. Reuse the logic from agenda_personal to get tasks and courses
    tareas_qs = Tarea.objects.filter(empleado=empleado_profile).order_by('fecha_limite')
    cursos_qs = EmpleadoCurso.objects.filter(empleado=empleado_profile).select_related('curso').order_by('curso__nombre')

    tareas_pendientes = [t for t in tareas_qs if not t.completada]
    tareas_historial = [t for t in tareas_qs if t.completada]
    
    cursos_pendientes = [c for c in cursos_qs if c.estado != 'APROBADO']
    cursos_historial = [c for c in cursos_qs if c.estado == 'APROBADO']

    # Cálculos para la vista (Admin viendo Empleado)
    total_asignados = len(cursos_qs)
    progreso_cumplimiento = 0
    if total_asignados > 0:
        progreso_cumplimiento = int((len(cursos_historial) / total_asignados) * 100)
        
    total_tareas = len(tareas_qs)
    progreso_tareas = 0
    if total_tareas > 0:
        progreso_tareas = int((len(tareas_historial) / total_tareas) * 100)

    # 4. Prepare context for the template
    context = {
        **ctx, # Unpack the common context
        'tareas_pendientes': tareas_pendientes,
        'tareas_historial': tareas_historial,
        'cursos_pendientes': cursos_pendientes,
        'cursos_historial': cursos_historial,
        'tareas': tareas_qs, # Modals
        # 'usuario': admin_profile, # Already in ctx
        'empleado': empleado_profile,   # The employee being viewed (overrides ctx['empleado'] if different)
        # 'nombre_rol': "Administrador", # Already in ctx
        'is_admin_view': True,          # A flag to adjust the template for admin viewing
        # 'inventario': 1, # Already in ctx (as boolean)
        # 'usuarios': 1, # Already in ctx (as boolean)
        'today': timezone.now().date(),
        'progreso_cumplimiento': progreso_cumplimiento,
        'total_cursos': total_asignados,
        'progreso_tareas': progreso_tareas,
        'total_tareas': total_tareas,
    }
    
    return render(request, 'plantillas/agenda.html', context)

@login_required
def exportar_matriz_excel(request):
    ctx = obtener_contexto_usuario(request)
    # Permission check
    if not request.user.is_superuser and not ctx['consultar']:
        messages.error(request, "No tienes permisos para exportar reportes.")
        return redirect('gestion_cursos')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Matriz de Cumplimiento'

    estado_filtro = request.GET.get('estado')
    curso_filtro = request.GET.get('curso')
    cargo_filtro = request.GET.get('cargo')
    
    seguimiento = EmpleadoCurso.objects.select_related('empleado', 'curso', 'empleado__id_cargo').all()
    if estado_filtro: seguimiento = seguimiento.filter(estado=estado_filtro)
    if curso_filtro: seguimiento = seguimiento.filter(curso__id_curso=curso_filtro)
    if cargo_filtro: seguimiento = seguimiento.filter(empleado__id_cargo_id=cargo_filtro)
    seguimiento = seguimiento.order_by('empleado__nombre', 'curso__nombre')

    headers = ['Empleado', 'Identificación', 'Cargo', 'Curso Obligatorio', 'Estado', 'Fecha de Vencimiento']
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for item in seguimiento:
        sheet.append([item.empleado.nombre, item.empleado.identificacion, item.empleado.id_cargo.nombre, item.curso.nombre, item.get_estado_display(), item.fecha_vencimiento.strftime('%Y-%m-%d') if item.fecha_vencimiento else 'N/A'])

    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']: sheet.column_dimensions[col_letter].autosize = True

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="matriz_cumplimiento.xlsx"'
    workbook.save(response)
    return response

@login_required
def descargar_pdf_empleados(request):
    ctx = obtener_contexto_usuario(request)
    # Permission check
    if not request.user.is_superuser and not ctx['consultar']:
        messages.error(request, "No tienes permisos para descargar este reporte.")
        return redirect('empleados')

    # Obtener empleados (podrías aplicar los mismos filtros que en la lista si lo deseas)
    empleados = Empleado.objects.all().order_by('nombre')
    
    # Contexto para la plantilla PDF
    context = {
        **ctx, # Unpack the common context
        'empleados': empleados,
        'fecha': timezone.now(),
        'titulo': 'Listado de Empleados - Petrocentro'
    }
    
    # Renderizar plantilla y crear PDF
    template_path = 'dash/lista_empleados_pdf.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="listado_empleados.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)
    return response

@login_required
def exportar_empleados_excel(request):
    ctx = obtener_contexto_usuario(request)
    # Permission check
    if not request.user.is_superuser and not ctx['consultar']:
        messages.error(request, "No tienes permisos para exportar reportes.")
        return redirect('empleados')

    # Filtros
    busqueda = request.GET.get("buscar") 
    estado = request.GET.get("estado")
    filtro_cargo = request.GET.get("cargo")
    filtro_ubicacion = request.GET.get("ubicacion")

    lista_empleados = Empleado.objects.select_related('id_cargo', 'id_rol', 'id_ubicacion', 'estado').all()

    if estado and estado != "0":
        lista_empleados = lista_empleados.filter(estado=estado)
    if filtro_cargo and filtro_cargo != '0':
        lista_empleados = lista_empleados.filter(id_cargo_id=filtro_cargo)
    if filtro_ubicacion and filtro_ubicacion != '0':
        lista_empleados = lista_empleados.filter(id_ubicacion_id=filtro_ubicacion)
    if busqueda:
        lista_empleados = lista_empleados.filter(
            Q(nombre__icontains=busqueda) | Q(identificacion__icontains=busqueda)
        )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Empleados'

    headers = ['ID', 'Nombre', 'Identificación', 'Correo', 'Teléfono', 'Cargo', 'Ubicación', 'Rol', 'Estado', 'Ingreso']
    sheet.append(headers)
    for cell in sheet[1]: cell.font = Font(bold=True)

    for e in lista_empleados:
        rol_nombre = e.id_rol.nombre if e.id_rol else 'Sin Rol'
        rol = rol_nombre
        estado_nombre = e.estado.nombre if e.estado else 'N/A'
        sheet.append([e.id, e.nombre, e.identificacion, e.correo, e.telefono, e.id_cargo.nombre, e.id_ubicacion.nombre, rol, estado_nombre, e.fecha_ingreso])

    for col in ['B', 'D', 'F', 'G']: sheet.column_dimensions[col].autosize = True

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="empleados_filtrados.xlsx"'
    workbook.save(response)
    return response

@login_required
def asignar_curso_empleado(request, id_empleado):
    ctx = obtener_contexto_usuario(request)
    # Permission check
    if not request.user.is_superuser and not ctx['crear']:
        messages.error(request, "No tienes permisos para asignar cursos.")
        return redirect('empleados')

    if request.method == 'POST':
        empleado = get_object_or_404(Empleado, id=id_empleado)
        curso_id = request.POST.get('curso')
        
        if curso_id:
            curso = get_object_or_404(Curso, id_curso=curso_id)
            obj, created = EmpleadoCurso.objects.get_or_create(empleado=empleado, curso=curso)
            if created:
                messages.success(request, f"Curso '{curso.nombre}' asignado a {empleado.nombre}.")
                # Notificación por correo
                try:
                    send_mail(
                        'Nuevo Curso Asignado - Petrocentro',
                        f'Hola {empleado.nombre},\n\nSe te ha asignado el curso obligatorio "{curso.nombre}".\nPor favor ingresa a la plataforma para gestionarlo.\n\nAtentamente,\nEquipo de Formación.',
                        settings.EMAIL_HOST_USER,
                        [empleado.correo],
                        fail_silently=True
                    )
                except:
                    pass
                # Log
                LogActividad.objects.create(usuario_afectado=empleado, admin_responsable=ctx['usuario'], accion="Asignación Manual Curso", detalles=f"Curso: {curso.nombre}")
            else:
                messages.warning(request, f"El empleado ya tiene asignado el curso '{curso.nombre}'.")
    
    return redirect('empleados')

@login_required
def enviar_notificacion_empleado(request, id_empleado):
    ctx = obtener_contexto_usuario(request)
    # Permission check
    if not request.user.is_superuser and not ctx['crear']: # Assuming 'crear' or a specific 'notificar' permission
        messages.error(request, "No tienes permisos para enviar notificaciones.")
        return redirect('empleados')

    if request.method == 'POST':
        empleado = get_object_or_404(Empleado, id=id_empleado)
        asunto = request.POST.get('asunto')
        mensaje = request.POST.get('mensaje')
        
        try:
            send_mail(
                f"Notificación Petrocentro: {asunto}",
                f"Hola {empleado.nombre},\n\n{mensaje}\n\nAtentamente,\nAdministración.",
                settings.EMAIL_HOST_USER,
                [empleado.correo],
                fail_silently=False
            )
            messages.success(request, f"Correo enviado a {empleado.nombre}.")
            
            LogActividad.objects.create(usuario_afectado=empleado, admin_responsable=ctx['usuario'], accion="Notificación Enviada", detalles=f"Asunto: {asunto}")
        except Exception as e:
            messages.error(request, f"Error al enviar correo: {e}")
            
    return redirect('empleados')

@login_required
def crear_tarea_rapida_empleado(request, id_empleado):
    ctx = obtener_contexto_usuario(request)
    # Permission check
    if not request.user.is_superuser and not ctx['crear']: # Assuming 'crear' permission for tasks
        messages.error(request, "No tienes permisos para crear tareas.")
        return redirect('empleados')

    if request.method == 'POST':
        empleado = get_object_or_404(Empleado, id=id_empleado)
        titulo = request.POST.get('titulo')
        fecha = request.POST.get('fecha_limite')
        desc = request.POST.get('descripcion')
        
        Tarea.objects.create(
            titulo=titulo,
            descripcion=desc,
            fecha_limite=fecha,
            id_cargo=empleado.id_cargo,
            empleado=empleado
        )
        # Notificación por correo
        try:
            send_mail(
                'Nueva Tarea Asignada - Petrocentro',
                f'Hola {empleado.nombre},\n\nSe te ha asignado una nueva tarea: "{titulo}".\nFecha límite: {fecha}\n\nDescripción: {desc}\n\nPor favor ingresa a tu agenda para gestionarla.',
                settings.EMAIL_HOST_USER,
                [empleado.correo],
                fail_silently=True
            )
        except:
            pass
        messages.success(request, f"Tarea asignada a {empleado.nombre}.")
        
    return redirect('empleados')

@login_required
def descargar_pdf_cumplimiento(request):
    ctx = obtener_contexto_usuario(request)
    # Verificación de permisos (Admin o Supervisor)
    if not request.user.is_superuser and not ctx['consultar']:
        messages.error(request, "No tienes permisos para descargar este reporte.")
        return redirect('gestion_cursos')

    # Obtener datos de cumplimiento
    cursos_asignados = EmpleadoCurso.objects.select_related('empleado', 'curso', 'empleado__id_cargo').all().order_by('empleado__id_cargo__nombre', 'empleado__nombre')
    
    # Calculos básicos para el reporte
    total = cursos_asignados.count()
    aprobados = cursos_asignados.filter(estado='APROBADO').count()
    porcentaje = (aprobados / total * 100) if total > 0 else 0

    context = {
        **ctx, # Unpack the common context
        'registros': cursos_asignados,
        'fecha': timezone.now(),
        'titulo': 'Reporte de Cumplimiento de Formación',
        'resumen': {'total': total, 'aprobados': aprobados, 'porcentaje': round(porcentaje, 1)}
    }
    
    template_path = 'dash/reporte_cumplimiento_pdf.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_cumplimiento.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)
    return response

@login_required
def descargar_pdf_usuarios(request):
    ctx = obtener_contexto_usuario(request)
    # Permission check
    if not request.user.is_superuser and not ctx['consultar']:
        messages.error(request, "No tienes permisos para descargar este reporte.")
        return redirect('usuarios')

    usuarios = Usuario.objects.all().order_by('nombre')
    
    context = {
        **ctx, # Unpack the common context
        'usuarios': usuarios,
        'fecha': timezone.now(),
        'titulo': 'Reporte de Usuarios del Sistema'
    }
    
    template_path = 'dash/lista_usuarios_pdf.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="listado_usuarios.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)
    return response

@login_required
def notificar_vencimiento_curso(request, id_empleado_curso):
    ctx = obtener_contexto_usuario(request)
    # Verificar permisos básicos (Admin o Supervisor)
    if not request.user.is_superuser and not ctx['editar']: # Assuming 'editar' or a specific 'notificar' permission
        messages.error(request, "No tienes permisos para enviar notificaciones.")
        return redirect('gestion_cursos')

    item = get_object_or_404(EmpleadoCurso, id_empleado_curso=id_empleado_curso)
    
    try:
        asunto = f"Alerta de Vencimiento: {item.curso.nombre}"
        mensaje = f"Hola {item.empleado.nombre},\n\nEste es un aviso automático para informarle que su certificado del curso obligatorio '{item.curso.nombre}' está próximo a vencer el día {item.fecha_vencimiento}.\n\nPor favor, gestione su renovación lo antes posible para mantener su cumplimiento al día.\n\nAtentamente,\nGestión Humana y HSEQ."
        
        send_mail(asunto, mensaje, settings.EMAIL_HOST_USER, [item.empleado.correo], fail_silently=False)
        
        LogActividad.objects.create(usuario_afectado=item.empleado, admin_responsable=ctx['usuario'], accion="Recordatorio Vencimiento", detalles=f"Curso: {item.curso.nombre}")
        messages.success(request, f"Recordatorio de vencimiento enviado a {item.empleado.nombre}.")
    except Exception as e:
        messages.error(request, f"Error al enviar correo: {e}")

    return redirect('gestion_cursos')

@login_required
def notificar_vencimiento_masivo(request):
    ctx = obtener_contexto_usuario(request)
    if not request.user.is_superuser and not ctx['crear']: # Assuming 'crear' or a specific 'notificar' permission
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect('gestion_cursos')

    fecha_limite = timezone.now().date() + timedelta(days=30)
    pendientes = EmpleadoCurso.objects.filter(estado='APROBADO', fecha_vencimiento__lte=fecha_limite).select_related('empleado', 'curso')

    count = 0
    for item in pendientes:
        try:
            send_mail(
                f"Aviso de Vencimiento: {item.curso.nombre}",
                f"Hola {item.empleado.nombre},\n\nTu certificado del curso '{item.curso.nombre}' vence el {item.fecha_vencimiento}. Por favor gestiona su renovación.\n\nAtentamente,\nAdministración.",
                settings.EMAIL_HOST_USER,
                [item.empleado.correo],
                fail_silently=True
            )
            count += 1
        except: pass
    
    if count > 0:
        messages.success(request, f"Se han enviado {count} correos de recordatorio masivo.")
    else:
        messages.info(request, "No se enviaron correos (lista vacía o errores).")
        
    return redirect('gestion_cursos')

# -------------------------------- API CHAT INTERNO --------------------------------

@login_required
def chat_api(request):
    """Maneja el envío y recepción de mensajes vía AJAX"""
    ctx = obtener_contexto_usuario(request)
    usuario = ctx['usuario']
    
    if not usuario:
        return JsonResponse({'status': 'error', 'message': 'Usuario no identificado'}, status=404)

    if request.method == 'POST':
        mensaje = request.POST.get('mensaje')
        if mensaje:
            es_admin = request.user.is_superuser
            MensajeChat.objects.create(usuario=usuario, mensaje=mensaje, es_respuesta_admin=es_admin)
            return JsonResponse({'status': 'ok'})
    
    # GET: Obtener últimos 50 mensajes (Chat Global Simplificado para Demo)
    if request.user.is_superuser:
        mensajes = MensajeChat.objects.select_related('usuario').order_by('-fecha')[:50]
    else:
        # El empleado ve sus mensajes y las respuestas globales de admin
        mensajes = MensajeChat.objects.filter(Q(usuario=usuario) | Q(es_respuesta_admin=True)).select_related('usuario').order_by('-fecha')[:50]
        
    data = []
    for m in reversed(list(mensajes)):
        data.append({
            'user': m.usuario.nombre,
            'text': m.mensaje,
            'is_me': m.usuario.id == usuario.id,
            'time': m.fecha.strftime('%H:%M'),
            'es_admin': m.es_respuesta_admin
        })
    
    return JsonResponse({'mensajes': data})

# -------------------------------- PANELES DE AUDITORÍA Y RRHH --------------------------------

@login_required
def panel_auditoria(request):
    ctx = obtener_contexto_usuario(request)
    if not request.user.is_superuser:
        messages.error(request, "Acceso restringido a Administradores.")
        return redirect('index')
        
    logs = LogActividad.objects.select_related('usuario_afectado', 'admin_responsable').all().order_by('-fecha')[:200]
    
    context = {
        **ctx, # Unpack the common context
        'logs': logs,
        # 'usuario': usuario_profile, # Already in ctx
        # 'nombre_rol': "Administrador", # Already in ctx
        # 'usuarios': 1, # Already in ctx (as boolean)
        # 'inventario': 1, # Already in ctx (as boolean)
    }
    return render(request, 'dash/auditoria.html', context)

@login_required
def reporte_rrhh(request):
    ctx = obtener_contexto_usuario(request)
    if not request.user.is_superuser:
        messages.error(request, "Acceso restringido.")
        return redirect('index')

    # Optimizamos reporte RRHH con anotaciones masivas en una sola consulta
    empleados = Empleado.objects.select_related('id_cargo', 'id_ubicacion').annotate(
        total_c=Count('cursos_asignados', distinct=True),
        aprob_c=Count('cursos_asignados', filter=Q(cursos_asignados__estado='APROBADO'), distinct=True),
        total_t=Count('tarea', distinct=True),
        aprob_t=Count('tarea', filter=Q(tarea__completada=True), distinct=True)
    )
    reporte = []
    
    for emp in empleados:
        # Usamos los datos ya calculados por la base de datos
        perc_c = (emp.aprob_c / emp.total_c * 100) if emp.total_c > 0 else 0
        perc_t = (emp.aprob_t / emp.total_t * 100) if emp.total_t > 0 else 0
        
        reporte.append({'empleado': emp, 'curso_pct': round(perc_c,1), 'tarea_pct': round(perc_t,1), 'promedio': round((perc_c+perc_t)/2, 1)})
    
    context = {
        **ctx, # Unpack the common context
        'reporte': reporte,
        # 'usuario': usuario_profile, # Already in ctx
        # 'nombre_rol': "Administrador", # Already in ctx
        # 'inventario': 1, # Already in ctx (as boolean)
        # 'usuarios': 1 # Already in ctx (as boolean)
    }
    return render(request, 'dash/rrhh_reporte.html', context)

@login_required
def gestion_pqrs(request):
    """Vista para que el Administrador gestione todas las PQRS."""
    ctx = obtener_contexto_usuario(request)
    from paginaPetrocentro.models import PQRS # Import locally to break circular dependency
    
    # Verificación de Rol Administrador
    if not request.user.is_superuser and ctx['nombre_rol'] != "ADMINISTRADOR":
        messages.error(request, "No tienes permisos para este módulo.")
        return redirect('index')

    pqrs_qs = PQRS.objects.all().order_by('-fecha_creacion')
    
    # Filtros simples
    estado_f = request.GET.get('estado')
    if estado_f:
        pqrs_qs = pqrs_qs.filter(estado=estado_f)

    p = Paginator(pqrs_qs, 10)
    pagina = p.get_page(request.GET.get('page'))

    context = {
        **ctx,
        'pqrs_list': pagina,
        'estados': PQRS.ESTADOS,
    }
    return render(request, 'dash/pqrs_admin.html', context)

@login_required
def responder_pqrs(request, id_pqr):
    """Procesa el cambio de estado y respuesta de una PQRS."""
    ctx = obtener_contexto_usuario(request)
    from paginaPetrocentro.models import PQRS # Import locally
    
    # Verificación de Rol Administrador (Seguridad)
    if not request.user.is_superuser and ctx['nombre_rol'] != "ADMINISTRADOR":
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect('index')

    if request.method == 'POST':
        pqr = get_object_or_404(PQRS, id=id_pqr)
        pqr.respuesta = request.POST.get('respuesta')
        pqr.estado = request.POST.get('estado')

        # Notificación automática al usuario si se resuelve
        if pqr.estado == 'Resuelto' and pqr.respuesta:
            subject = f"Respuesta a su solicitud - Radicado {pqr.radicado}"
            message = f"Hola {pqr.nombre},\n\nHemos dado respuesta a su {pqr.tipo}.\n\nRespuesta:\n{pqr.respuesta}"
            try:
                send_mail(subject, message, settings.EMAIL_HOST_USER, [pqr.correo], fail_silently=True)
            except: pass
        
        pqr.save()
        messages.success(request, f"Radicado {pqr.radicado} actualizado correctamente.")
    return redirect('gestion_pqrs')

@login_required
def gestion_cotizaciones(request):
    """Vista para que el Administrador gestione todas las solicitudes de cotización."""
    ctx = obtener_contexto_usuario(request)
    from paginaPetrocentro.models import Cotizacion # Import locally
    
    if not request.user.is_superuser and ctx['nombre_rol'] != "ADMINISTRADOR":
        messages.error(request, "No tienes permisos para este módulo.")
        return redirect('index')

    cotizaciones_qs = Cotizacion.objects.all().order_by('-fecha_creacion')
    
    p = Paginator(cotizaciones_qs, 15)
    pagina = p.get_page(request.GET.get('page'))

    context = {
        **ctx,
        'cotizaciones_list': pagina,
        'total_cotizaciones': cotizaciones_qs.count(),
    }
    return render(request, 'dash/cotizaciones_admin.html', context)

# --- NUEVO MÓDULO DE INVENTARIO ---

@login_required
def gestion_inventario(request):
    """
    Vista para listar los equipos del inventario (basada en el nuevo modelo InventoryItem).
    """
    ctx = obtener_contexto_usuario(request)

    # Verificación robusta de permisos
    tiene_permiso = ctx.get('inventario') or request.user.is_superuser
    
    if not tiene_permiso:
        messages.error(request, "No tienes permisos para acceder al módulo de inventario.")
        return redirect('index')

    # Sedes autorizadas según requerimiento operativo (Definidas para todos los roles)
    VALID_BASES = ["Chichimene", "Mocoa", "Sibaté"]

    # --- OPTIMIZACIÓN: Solo ejecutar mantenimiento pesado si se solicita expresamente ---
    # Esto evita que la página tarde segundos en cargar en cada visita del administrador.
    ejecutar_mantenimiento = request.GET.get('maint') == '1'
    if request.user.is_superuser and ejecutar_mantenimiento:
        # --- MANTENIMIENTO DE ROLES CRÍTICOS ---
        if not Rol.objects.filter(nombre="Supervisor").exists():
            Rol.objects.get_or_create(nombre="Supervisor")

        # --- MANTENIMIENTO DE SEDES/UBICACIONES ---
        existing_bases = list(Ubicacion.objects.filter(nombre__in=VALID_BASES).values_list('nombre', flat=True))
        if len(existing_bases) < len(VALID_BASES):
            for b_name in VALID_BASES:
                if b_name not in existing_bases:
                    Ubicacion.objects.get_or_create(nombre=b_name)
        
        # Re-asignar recursos de sedes no autorizadas solo si existen
        bad_bases = Ubicacion.objects.exclude(nombre__in=VALID_BASES)
        if bad_bases.exists():
            fallback_base = Ubicacion.objects.filter(nombre="Chichimene").first()
            if fallback_base:
                Empleado.objects.filter(id_ubicacion__in=bad_bases).update(id_ubicacion=fallback_base)
                InventoryItem.objects.filter(base__in=bad_bases).update(base=fallback_base)
                bad_bases.delete()

        # --- MANTENIMIENTO PERMANENTE DE CATEGORÍAS ---
        WHITELIST = [
            "Proceso Presurizado", "Proceso atmosférico", "Inyección Bombeo y Compresión",
            "Control presión en Superficie", "Quemadores Gas", "Tubería y Accesorios",
            "Cargaderos", "Tableros de control y Distribución", "Generación e Iluminación",
            "Laboratorio y Metrología", "Medición Proceso", "HSE", "Campamento",
            "Tecnología", "Materiales de Instrumentación", "Herramienta", "Ferreteria"
        ]
        existing_cats = list(InventoryCategory.objects.filter(name__in=WHITELIST).values_list('name', flat=True))
        if len(existing_cats) < len(WHITELIST):
            for cat_name in WHITELIST:
                if cat_name not in existing_cats:
                    InventoryCategory.objects.get_or_create(name=cat_name)

        # Re-asignar items en categorías no autorizadas solo si existen
        bad_categories = InventoryCategory.objects.exclude(name__in=WHITELIST)
        if bad_categories.exists():
            ferreteria_cat = InventoryCategory.objects.filter(name="Ferreteria").first()
            if ferreteria_cat:
                InventoryItem.objects.filter(category__in=bad_categories).update(category=ferreteria_cat)
                bad_categories.delete()

        def sugerir_categoria(nombre_equipo):
            n = str(nombre_equipo).upper()
            if any(x in n for x in ["TANQUE", "TK", "BARRIL", "BLS", "VASA", "PISCINA"]): return "Proceso atmosférico"
            if any(x in n for x in ["BOMBA", "MOTOR", "COMPRE", "INYEC", "TRIPLEX", "UNIDAD INY", "CENTRIFUGA", "ROPER"]): return "Inyección Bombeo y Compresión"
            if any(x in n for x in ["VALVULA", "CODO", "TUBO", "ACCESORIO", "MANGUERA", "BRIDA", "UNION", "NIPLE", "TEE", "REDUCCION"]): return "Tubería y Accesorios"
            if any(x in n for x in ["GENERADOR", "PLANTA", "REFLECTOR", "LAMPARA", "ILUM", "EXTENSION"]): return "Generación e Iluminación"
            if any(x in n for x in ["TABLERO", "BREAKER", "CABLE", "ELECTRICO", "CONTACTOR", "TRANSFORMADOR"]): return "Tableros de control y Distribución"
            if any(x in n for x in ["TERMOMETRO", "MEDIDOR", "CAUDALIMETRO", "FLUJOM", "INSTRUM", "MANOMETRO", "SENSOR"]): return "Medición Proceso"
            if any(x in n for x in ["CASCO", "EXTINTOR", "BOTIQUIN", "HSE", "SEGURIDAD", "CONO", "ARNES", "GUANTE", "GAFA", "BOTA"]): return "HSE"
            if any(x in n for x in ["COMPUTADOR", "MONITOR", "TECLADO", "IMPRESORA", "PORTATIL", "TABLET", "CAMARA", "RADIO", "UPS"]): return "Tecnología"
            if any(x in n for x in ["HERRAMIENTA", "LLAVE", "DESTORNILLADOR", "MARTILLO", "TALADRO", "PULIDORA", "PRENSA"]): return "Herramienta"
            if any(x in n for x in ["TUBO", "CASING", "VARILLA", "TUBERIA"]): return "Tubería y Accesorios"
            if any(x in n for x in ["MESA", "SILLA", "CAMA", "AIRE", "NEVERA", "ESTUFA", "MICROONDAS"]): return "Campamento"
            if any(x in n for x in ["SEPARADOR", "CHIMENEA", "TEA", "SCRUBBER"]): return "Proceso Presurizado"
            return None

        # --- LIMPIEZA DE EQUIPOS FANTASMA (Encabezados importados como datos) ---
        trash_words = ["TIPO DE EO", "CATEGORÍA", "DESCRIPCIÓN DEL ARTÍCULO", "ESPECIFICACIÓN TÉCNICA", "MARCA", "SERIAL", "TAG", "UNIDAD", "PROPIEDAD"]
        trash_qs = InventoryItem.objects.filter(
            Q(name__icontains="CATEGORÍA") | Q(name__icontains="DESCRIPCIÓN") | 
            Q(name__icontains="ARTÍCULOS") | Q(sku__in=["TAG", "SERIAL", "SKU", "None"])
        )
        trash_count = trash_qs.count()
        if trash_count > 0: trash_qs.delete()

        # --- MANTENIMIENTO DE DATOS (Descripciones Estandarizadas [NOMBRE] [ESPEC] / [MARCA]) ---
        # OPTIMIZACIÓN: Solo procesamos items que NO tengan descripción o categoría asignada
        items_limpieza = InventoryItem.objects.filter(
            Q(description__isnull=True) | Q(description="") | Q(category__isnull=True)
        )
        
        cambios = []
        for item in items_limpieza:
            modificado = False
            
            n_limpio = str(item.name).strip().upper()
            e_limpio = str(item.technical_spec).strip().upper() if item.technical_spec else ""
            m_limpio = str(item.brand).strip().upper() if item.brand else ""
            
            if e_limpio and e_limpio in n_limpio: e_limpio = ""
            base_txt = n_limpio
            if e_limpio and e_limpio not in ["NONE", "-", "N/A", ""]:
                base_txt = f"{base_txt} {e_limpio}"
            
            if m_limpio and m_limpio not in ["NONE", "-", "PETROCENTRO", "N/A", ""] and m_limpio not in n_limpio:
                desc_estandar = f"{base_txt} / {m_limpio}"
            else:
                desc_estandar = base_txt

            desc_estandar = desc_estandar.replace("  ", " ").strip()
            
            if item.description != desc_estandar:
                item.description = desc_estandar
                modificado = True

            # 2. Recategorización dinámica (Existente)
            nombre_cat_sugerida = sugerir_categoria(item.description or item.name)
            if nombre_cat_sugerida:
                cat_obj = InventoryCategory.objects.get(name=nombre_cat_sugerida)
                if item.category != cat_obj:
                    item.category = cat_obj
                    modificado = True

            if modificado:
                cambios.append(item)
        
        if cambios:
            InventoryItem.objects.bulk_update(cambios, ['category', 'description'])
            msg_maint = f"Mantenimiento: Se estandarizaron {len(cambios)} descripciones técnica."
            if trash_count > 0: msg_maint += f" Se eliminaron {trash_count} equipos inválidos."
            messages.success(request, msg_maint)

    items_qs = InventoryItem.objects.select_related('category', 'base', 'supplier', 'project').all()

    # --- RESUMEN POR SEDE (INDEPENDIENTE) ---
    # Calculamos métricas en tiempo real por cada una de las 3 bases operativas
    base_inventory_summary = Ubicacion.objects.filter(
        nombre__in=VALID_BASES
    ).annotate(
        total_items=Count('inventoryitem', distinct=True),
        total_units=Sum('inventoryitem__current_stock'),
        total_value=Sum(F('inventoryitem__current_stock') * F('inventoryitem__unit_price'))
    ).order_by('nombre')

    # --- CÁLCULO DE STOCK GLOBAL (GENERAL) ---
    # Agrupamos en Python para incluir IDs de items para selección directa desde el Dashboard (Funcionalidad de Carrito)
    items_for_summary = InventoryItem.objects.filter(base__nombre__in=VALID_BASES).select_related('category', 'base')
    summary_map = {}
    for it in items_for_summary:
        key = it.description or it.name
        if key not in summary_map:
            summary_map[key] = {
                'description': key,
                'category__name': it.category.name if it.category else "Sin Categoría",
                'total_empresa': 0,
                'stock_chichimene': 0, 'id_chichimene': None,
                'stock_mocoa': 0, 'id_mocoa': None,
                'stock_sibate': 0, 'id_sibate': None,
            }
        
        summary_map[key]['total_empresa'] += it.current_stock
        b_norm = it.base.nombre.upper() if it.base else ""
        if "CHICHIMENE" in b_norm:
            summary_map[key]['stock_chichimene'] += it.current_stock
            if it.current_stock > 0 and not summary_map[key]['id_chichimene']: summary_map[key]['id_chichimene'] = it.id
        elif "MOCOA" in b_norm:
            summary_map[key]['stock_mocoa'] += it.current_stock
            if it.current_stock > 0 and not summary_map[key]['id_mocoa']: summary_map[key]['id_mocoa'] = it.id
        elif "SIBATE" in b_norm or "SIBATÉ" in b_norm:
            summary_map[key]['stock_sibate'] += it.current_stock
            if it.current_stock > 0 and not summary_map[key]['id_sibate']: summary_map[key]['id_sibate'] = it.id
    
    global_stock_summary = sorted(summary_map.values(), key=lambda x: x['description'])

    category_summary = InventoryCategory.objects.annotate(
        total_tipos=Count('items', distinct=True),
        total_unidades=Sum('items__current_stock')
    ).order_by('name')

    # --- LÓGICA DE AGRUPACIÓN DE SOLICITUDES (EXPERIENCIA DE SOLICITUD ÚNICA) ---
    def get_grouped_requests(queryset):
        grouped = {}
        for req in queryset:
            # Si no tiene batch_id (registros antiguos), creamos uno virtual para agruparlo solo a él
            key = req.batch_id if req.batch_id else f"OLD_{req.id}"
            if key not in grouped:
                grouped[key] = {
                    'id': req.id, 
                    'batch_id': req.batch_id, 
                    'supervisor': req.supervisor,
                    'dest': req.destination_base.nombre if req.destination_base else (req.destination_project.name if req.destination_project else "Proyecto"),
                    'reason': req.reason, 
                    'date': req.created_at,
                    'status': req.status,
                    'processed_by': req.processed_by,
                    'processed_at': req.processed_at,
                    'requested_items': []
                }
            grouped[key]['requested_items'].append(req)
        return grouped

    # 1. Solicitudes Pendientes (Vista Admin)
    if ctx['is_admin'] or request.user.is_superuser: # Administradores ven todas
        raw_pending = TransferRequest.objects.filter(status='PENDING').select_related('item', 'item__base', 'destination_base', 'supervisor').order_by('-created_at')
    elif ctx['is_supervisor']: # Supervisores ven las que ellos mismos crearon
        raw_pending = TransferRequest.objects.filter(status='PENDING', supervisor=ctx['usuario']).select_related('item', 'item__base', 'destination_base', 'supervisor').order_by('-created_at')
    else:
        raw_pending = TransferRequest.objects.none()
    grouped_pending = get_grouped_requests(raw_pending)

    # Historial de solicitudes (Procesadas)
    if ctx['is_admin'] or request.user.is_superuser: # Historial global para admin
        raw_history = TransferRequest.objects.exclude(status='PENDING').select_related('item', 'item__base', 'destination_base', 'supervisor', 'processed_by').order_by('-processed_at')
    elif ctx['is_supervisor']: # El supervisor ve lo que pidió Y lo que recibió en su base
        history_filter = Q(supervisor=ctx['usuario'])
        if ctx['empleado'] and ctx['empleado'].id_ubicacion:
            history_filter |= Q(destination_base=ctx['empleado'].id_ubicacion)
        raw_history = TransferRequest.objects.filter(history_filter).exclude(status='PENDING').select_related('item', 'item__base', 'destination_base', 'supervisor', 'processed_by').order_by('-processed_at')
    else:
        raw_history = TransferRequest.objects.none()
    grouped_history = get_grouped_requests(raw_history)

    # 3. Solicitudes Entrantes (Pendientes por recibir)
    # Los Admins ven todo lo que está en tránsito. Supervisores solo lo que va a su base.
    if ctx['is_admin'] or request.user.is_superuser:
        raw_incoming = TransferRequest.objects.filter(status='SENT').select_related('item', 'item__base', 'supervisor', 'destination_base', 'destination_project')
    elif ctx['empleado'] and ctx['empleado'].id_ubicacion:
        raw_incoming = TransferRequest.objects.filter(destination_base=ctx['empleado'].id_ubicacion, status='SENT').select_related('item', 'item__base', 'supervisor', 'destination_project')
    else:
        raw_incoming = TransferRequest.objects.none()
    grouped_incoming = get_grouped_requests(raw_incoming)

    # --- MOVIMIENTOS ESPECÍFICOS PARA TABS SOLICITADAS ---
    # 1. Entradas y Salidas (E/S)
    entries_exits = StockTransaction.objects.filter(
        movement_type__in=['ENTRY', 'EXIT', 'RETURN', 'ADJUST_IN', 'ADJUST_OUT']
    ).select_related('item', 'responsible', 'item__base').order_by('-date')[:20]
    # 2. Control de Cambios (Traslados entre sedes)
    change_control = StockTransaction.objects.filter(movement_type='TRANSFER').select_related('item', 'responsible', 'origin_base', 'destination_base').order_by('-date')[:20]

    # --- FILTROS ---
    search_query = request.GET.get('search')
    category_filter = request.GET.get('category')
    project_filter = request.GET.get('project')
    base_filter = request.GET.get('base')
    critical_filter = request.GET.get('critical')
    supplier_filter = request.GET.get('supplier')
    owner_filter = request.GET.get('owner')
    status_filter = request.GET.get('status')
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')

    if search_query:
        items_qs = items_qs.filter(
            Q(name__icontains=search_query) |
            Q(sku__icontains=search_query) |
            Q(serial__icontains=search_query) | # Añadido Serial a la búsqueda
            Q(tag__icontains=search_query) |     # Añadido TAG a la búsqueda
            Q(category__name__icontains=search_query)
        )
    if category_filter and category_filter != '0':
        items_qs = items_qs.filter(category__id=category_filter)
    if project_filter and project_filter != '0':
        items_qs = items_qs.filter(project_id=project_filter)
    if base_filter and base_filter != '0':
        items_qs = items_qs.filter(base_id=base_filter)
    if critical_filter == '1':
        items_qs = items_qs.filter(current_stock__lte=F('min_stock'))
    if supplier_filter and supplier_filter != '0':
        items_qs = items_qs.filter(supplier_id=supplier_filter)
    if owner_filter:
        items_qs = items_qs.filter(owner__icontains=owner_filter)

    if status_filter:
        items_qs = items_qs.filter(status=status_filter)

    # Filtros de fecha de creación
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            items_qs = items_qs.filter(created_at__date__gte=date_from)
        except ValueError:
            messages.error(request, "Formato de fecha 'Desde' inválido.")
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            items_qs = items_qs.filter(created_at__date__lte=date_to)
        except ValueError:
            messages.error(request, "Formato de fecha 'Hasta' inválido.")

    # --- PAGINACIÓN (Soluciona el problema de carga lenta) ---
    paginator = Paginator(items_qs, 50) # 50 items por página
    page_number = request.GET.get('page')
    items_page = paginator.get_page(page_number)

    # --- KPIs OPTIMIZADOS: Consolidamos múltiples consultas en una sola agregación condicional ---
    kpi_stats = items_qs.aggregate(
        total_val=Sum(F('current_stock') * F('unit_price')),
        total_stock=Sum('current_stock'),
        available=Sum(Case(When(status='OPERATIVO', then=F('current_stock')), default=0, output_field=IntegerField())),
        in_project=Sum(Case(When(status='EN PROYECTO', then=F('current_stock')), default=0, output_field=IntegerField())),
        in_maint=Sum(Case(When(status='FUERA DE SERVICIO', then=F('current_stock')), default=0, output_field=IntegerField())),
        critical=Count(Case(When(current_stock__lte=F('min_stock'), then=1))),
        consumables_crit=Count(Case(When(equipment_type='SUMINISTRO', current_stock__lte=F('min_stock'), then=1)))
    )

    total_inventory_value = kpi_stats['total_val'] or 0
    total_physical_units = kpi_stats['total_stock'] or 0
    available_units = kpi_stats['available'] or 0
    project_units = kpi_stats['in_project'] or 0
    maintenance_units = kpi_stats['in_maint'] or 0
    critical_items_count = kpi_stats['critical'] or 0
    consumables_critical_count = kpi_stats['consumables_crit'] or 0

    # Alertas de calibración (también basadas en items_qs filtrados)
    today_date = timezone.now().date()
    limit_date = today_date + timedelta(days=30) # Próximos 30 días
    calibration_alerts_count = items_qs.filter(
        calibration_date__lte=limit_date
    ).exclude(status__in=['VENDIDO', 'DADO DE BAJA']).count()
    
    # Alertas de mantenimiento técnico
    maintenance_alerts_count = items_qs.filter(
        maintenance_date__lte=limit_date
    ).exclude(status__in=['VENDIDO', 'DADO DE BAJA']).count()
 
    consumables_count = items_qs.filter(equipment_type='SUMINISTRO').count()

    # --- MÉTRICAS DASHBOARD ---
    last_30_days = timezone.now() - timedelta(days=30)
    top_moved = StockTransaction.objects.filter(
        movement_type='EXIT', date__gte=last_30_days
    ).values('item__name').annotate(total_qty=Sum('quantity')).order_by('-total_qty')[:5]
    faltantes_count = items_qs.filter(current_stock=0).count()
    movimientos_mes = StockTransaction.objects.filter(date__gte=last_30_days).count()
    
    # Comparativo de entradas y salidas en los últimos 30 días
    total_entries_30d = StockTransaction.objects.filter(
        date__gte=last_30_days, movement_type__in=['ENTRY', 'RETURN', 'ADJUST_IN']).aggregate(Sum('quantity'))['quantity__sum'] or 0
    total_exits_30d = StockTransaction.objects.filter(
        date__gte=last_30_days, movement_type__in=['EXIT', 'ADJUST_OUT']).aggregate(Sum('quantity'))['quantity__sum'] or 0

    # Notificaciones de solicitudes pendientes para el badge del menú
    notificaciones_inventario = 0
    if ctx['is_admin'] or request.user.is_superuser:
        notificaciones_inventario = TransferRequest.objects.filter(status='PENDING').count()

    # --- UNIFICACIÓN: LISTA DE SECCIÓN (PERSONAL) ---
    # Recuperamos el personal perteneciente a la sección/base seleccionada
    personal_qs = Empleado.objects.select_related('id_cargo', 'id_rol', 'id_ubicacion', 'estado').all().order_by('nombre')
    if base_filter and base_filter != '0' and base_filter.isdigit():
        personal_qs = personal_qs.filter(id_ubicacion_id=base_filter)
    
    paginator_personal = Paginator(personal_qs, 15)
    page_personal = request.GET.get('page_personal')
    personal_page = paginator_personal.get_page(page_personal)

    # --- DATOS PARA LA PESTAÑA DE REPORTES ADMINISTRATIVOS ---
    critical_items_list = InventoryItem.objects.filter(current_stock__lte=F('min_stock')).select_related('category', 'base').order_by('name')
    calibration_items_list = InventoryItem.objects.filter(calibration_date__lte=limit_date).exclude(status__in=['VENDIDO', 'DADO DE BAJA']).select_related('category', 'base').order_by('calibration_date')
    recent_transactions = StockTransaction.objects.select_related('item', 'responsible', 'origin_base', 'destination_base').order_by('-date')[:50] # Más movimientos para el reporte
    import_history = LogActividad.objects.filter(accion="Importación Masiva Inventario").select_related('admin_responsable').order_by('-fecha')[:10]
    context = {
        **ctx, # Incluye el contexto común para el menú
        'notificaciones_inventario': notificaciones_inventario,
        'usuarios_list': Usuario.objects.filter(estado_id=1).order_by('nombre'),
        'mov_types': StockTransaction.MOVEMENT_TYPES,
        'equipment_types': ['ACTIVO', 'ACCESORIO', 'SUMINISTRO', 'OTROS', 'VENDIDO'],
        'items': items_page, # Usamos el objeto paginado
        'personal_seccion': personal_page, # Listado unificado de empleados
        'global_summary': global_stock_summary,
        'base_inventory_summary': base_inventory_summary,
        'category_summary': category_summary, # Aseguramos que category_summary esté en el contexto
        'grouped_pending': grouped_pending,
        'grouped_history': grouped_history,
        'grouped_incoming': grouped_incoming,
        'entries_exits': entries_exits,
        'change_control': change_control,
        'all_items_modals': InventoryItem.objects.select_related('base', 'project', 'category').all(),
        'categories': InventoryCategory.objects.all().order_by('name'),
        'proveedores': Proveedor.objects.all().order_by('nombre'),
        'proyectos': Project.objects.all().order_by('name'),
        'bases': Ubicacion.objects.all(),
        'stock_movements': StockTransaction.objects.select_related('item', 'responsible').order_by('-date')[:10],
        'total_inventory_value': total_inventory_value,
        'critical_items_count': critical_items_count,
        'calibration_alerts_count': calibration_alerts_count,
        'available_units': available_units,
        'project_units': project_units,
        'maintenance_units': maintenance_units,
        'consumables_critical_count': consumables_critical_count,
        'maintenance_alerts_count': maintenance_alerts_count,
        'consumables_count': consumables_count,
        'total_physical_units': total_physical_units,
        'top_moved': top_moved,
        'faltantes_count': faltantes_count,
        'movimientos_mes': movimientos_mes,
        'total_entries_30d': total_entries_30d,
        'total_exits_30d': total_exits_30d,
        'critical_items_list': critical_items_list,
        'calibration_items_list': calibration_items_list,
        'recent_transactions': recent_transactions,
        'import_history': import_history,
        'title_page': "Inventario General",
        'now': timezone.now(),
        'filters': {
            'search': search_query,
            'category': category_filter,
            'project': project_filter,
            'base': base_filter,
            'critical': critical_filter,
            'supplier': supplier_filter,
            'owner': owner_filter,
            'status': status_filter,
            'date_from': date_from_str,
            'date_to': date_to_str,
        }
    }
    return render(request, 'dash/new_inventory.html', context)

@login_required
def exportar_inventario_excel(request):
    """Genera un reporte Excel del inventario actual valorizado para revisión administrativa."""
    ctx = obtener_contexto_usuario(request)
    if not request.user.is_superuser and not ctx['consultar']:
        messages.error(request, "No tienes permisos para exportar reportes.")
        return redirect('gestion_inventario')

    items = InventoryItem.objects.select_related('category', 'base', 'supplier').all().order_by('category__name', 'name')

    # APLICAR FILTROS (Paridad total con la vista de gestión para exportaciones precisas)
    search_query = request.GET.get('search')
    category_filter = request.GET.get('category')
    project_filter = request.GET.get('project')
    base_filter = request.GET.get('base')
    critical_filter = request.GET.get('critical')
    supplier_filter = request.GET.get('supplier')
    status_filter = request.GET.get('status')

    if search_query:
        items = items.filter(
            Q(name__icontains=search_query) |
            Q(sku__icontains=search_query) |
            Q(serial__icontains=search_query) |
            Q(tag__icontains=search_query) |
            Q(category__name__icontains=search_query)
        )
    if category_filter and category_filter != '0':
        items = items.filter(category__id=category_filter)
    if project_filter and project_filter != '0':
        items = items.filter(project_id=project_filter)
    if base_filter and base_filter != '0':
        items = items.filter(base_id=base_filter)
    if critical_filter == '1':
        items = items.filter(current_stock__lte=F('min_stock'))
    if supplier_filter and supplier_filter != '0':
        items = items.filter(supplier_id=supplier_filter)
    if status_filter:
        items = items.filter(status=status_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Valorizado"

    headers = [
        'SKU', 'Nombre', 'Categoría', 'Tipo', 'Marca', 'Serial', 'TAG', 
        'Base', 'Proyecto', 'Stock Actual', 'UOM', 'Precio Unitario', 'Valor Total', 'Estado'
    ]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for item in items:
        ws.append([
            item.sku, item.name, 
            item.category.name if item.category else 'N/A',
            item.equipment_type, item.brand, item.serial, item.tag,
            item.base.nombre if item.base else 'N/A',
            item.project.name if item.project else 'N/A',
            item.current_stock, item.unit_of_measure,
            float(item.unit_price), float(item.current_stock * item.unit_price),
            item.status
        ])

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario_valorizado_petrocentro.xlsx"'
    wb.save(response)
    return response

@login_required
def descargar_pdf_inventario_general(request):
    """Genera un reporte PDF ejecutivo consolidado por categoría."""
    ctx = obtener_contexto_usuario(request)
    if not request.user.is_superuser and not ctx['consultar']:
        messages.error(request, "No tienes permisos para descargar reportes.")
        return redirect('gestion_inventario')

    categories = InventoryCategory.objects.annotate(
        total_tipos=Count('items', distinct=True),
        total_unidades=Sum('items__current_stock'),
        valor_total=Sum(F('items__current_stock') * F('items__unit_price'))
    ).filter(total_tipos__gt=0).order_by('name')

    total_global = categories.aggregate(Sum('valor_total'))['valor_total__sum'] or 0

    context = {
        **ctx,
        'categories': categories,
        'total_global': total_global,
        'fecha': timezone.now(),
        'titulo': 'Reporte Ejecutivo de Inventario Valorizado'
    }
    
    template_path = 'dash/reporte_inventario_pdf.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario_gerencial.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err: return HttpResponse('Error al generar el PDF', status=500)
    return response

@login_required
def exportar_movimientos_excel(request):
    """Exporta el historial de movimientos para trazabilidad, opcionalmente filtrado por equipo (SKU)."""
    ctx = obtener_contexto_usuario(request)
    if not request.user.is_superuser and not ctx['consultar']:
        messages.error(request, "No tienes permisos para exportar reportes.")
        return redirect('gestion_inventario')

    # Obtener filtros extendidos para trazabilidad
    sku_filter = request.GET.get('sku')
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    responsible_id = request.GET.get('responsible')
    movement_type = request.GET.get('type')
    category_id = request.GET.get('category')
    equipment_type = request.GET.get('equipment_type')

    movimientos = StockTransaction.objects.select_related(
        'item', 'responsible', 'origin_base', 'destination_base', 'item__category'
    ).all().order_by('-date')
    
    if sku_filter:
        movimientos = movimientos.filter(item__sku=sku_filter)
    
    if date_from_str:
        try:
            movimientos = movimientos.filter(date__date__gte=datetime.strptime(date_from_str, '%Y-%m-%d').date())
        except ValueError: pass
        
    if date_to_str:
        try:
            movimientos = movimientos.filter(date__date__lte=datetime.strptime(date_to_str, '%Y-%m-%d').date())
        except ValueError: pass

    if responsible_id and responsible_id != '0':
        movimientos = movimientos.filter(responsible_id=responsible_id)
    
    if movement_type and movement_type != '0':
        movimientos = movimientos.filter(movement_type=movement_type)
        
    if category_id and category_id != '0':
        movimientos = movimientos.filter(item__category_id=category_id)
        
    if equipment_type and equipment_type != '0':
        movimientos = movimientos.filter(item__equipment_type=equipment_type)

    filename = "reporte_trazabilidad_filtrado.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trazabilidad de Stock"
    headers = ['Fecha', 'Equipo', 'SKU', 'Categoría', 'Tipo Recurso', 'Movimiento', 'Cantidad', 'Responsable', 'Origen (Sede/Proy)', 'Destino (Sede/Proy)', 'Motivo']
    ws.append(headers)
    for cell in ws[1]: cell.font = Font(bold=True)

    for m in movimientos:
        # Formatear origen y destino combinando Sede y Proyecto para claridad en reportes
        origen = f"{m.origin_base.nombre if m.origin_base else ''} {m.origin_project.name if m.origin_project else ''}".strip() or 'N/A'
        destino = f"{m.destination_base.nombre if m.destination_base else ''} {m.destination_project.name if m.destination_project else ''}".strip() or 'N/A'
        
        ws.append([
            m.date.strftime('%Y-%m-%d %H:%M'),
            m.item.name,
            m.item.sku,
            m.item.category.name if m.item.category else 'N/A',
            m.item.equipment_type or 'ACTIVO',
            m.get_movement_type_display(), m.quantity,
            m.responsible.nombre if m.responsible else 'Sistema',
            origen,
            destino,
            m.reason
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def reiniciar_inventario_data(request):
    """Borra físicamente toda la información de las tablas de inventario."""
    ctx = obtener_contexto_usuario(request)
    # Permitir a superusuarios o al rol ADMINISTRADOR
    if not request.user.is_superuser and ctx['nombre_rol'] != "ADMINISTRADOR":
        messages.error(request, "Acción no autorizada.")
        return redirect('gestion_inventario')
    
    try:
        with transaction.atomic():
            # Eliminación en cascada manual para asegurar limpieza total
            StockTransaction.objects.all().delete()
            TransferRequest.objects.all().delete()
            InventoryItem.objects.all().delete()
            InventoryCategory.objects.all().delete()
            Project.objects.all().delete()
            
            LogActividad.objects.create(
                admin_responsable=ctx['usuario'],
                accion="REINICIO TOTAL INVENTARIO",
                detalles="Se vaciaron todas las tablas. Se eliminaron los registros previos para carga limpia."
            )
        messages.success(request, "¡Inventario vaciado! Todos los equipos (incluidos los 692 previos) han sido eliminados exitosamente.")
    except Exception as e:
        messages.error(request, f"Error al limpiar: {e}")
        
    return redirect(f"{reverse('gestion_inventario')}?tab=sede")

@login_required
def importar_inventario_excel(request):
    """Procesa un archivo Excel para carga masiva de equipos del inventario."""
    ctx = obtener_contexto_usuario(request)
    # Verificación de permisos: Admin o Superusuario
    if not request.user.is_superuser and not ctx.get('crear'):
        messages.error(request, "No tienes permisos para realizar importaciones.")
        return redirect('gestion_inventario')

    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        if not archivo.name.endswith('.xlsx'):
            messages.error(request, "El archivo debe ser un libro de Excel (.xlsx)")
            return redirect(f"{reverse('gestion_inventario')}?tab=sede")

        try:
            wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True) # Mejorar rendimiento con read_only
            
            # Intentar buscar la hoja "inventario", si no, usar la primera hoja disponible
            if "inventario" in wb.sheetnames:
                sheet = wb["inventario"]
            else:
                sheet = wb.worksheets[0] # Toma la primera hoja por defecto (Hoja 1)
            
            creados = 0
            actualizados = 0

            # Lista oficial definitiva de categorías para el mantenimiento post-importación
            WHITELIST_CATEGORIES = [
                "Proceso Presurizado", "Proceso atmosférico", "Inyección Bombeo y Compresión",
                "Control presión en Superficie", "Quemadores Gas", "Tubería y Accesorios",
                "Cargaderos", "Tableros de control y Distribución", "Generación e Iluminación",
                "Laboratorio y Metrología", "Medición Proceso", "HSE", "Campamento",
                "Tecnología", "Materiales de Instrumentación", "Herramienta", "Ferreteria"
            ]

            def normalizar_categoria(nombre_excel):
                if not nombre_excel or str(nombre_excel).strip() == "": return None
                n = str(nombre_excel).strip().upper()
                
                mapping = {
                    "PRESUR": "Proceso Presurizado",
                    "ATMOSF": "Proceso atmosférico",
                    "INYEC": "Inyección Bombeo y Compresión", "BOMBE": "Inyección Bombeo y Compresión", "COMPR": "Inyección Bombeo y Compresión",
                    "CONTROL P": "Control presión en Superficie",
                    "QUEMAD": "Quemadores Gas",
                    "TUBER": "Tubería y Accesorios", "ACCES": "Tubería y Accesorios",
                    "CARGAD": "Cargaderos",
                    "TABLER": "Tableros de control y Distribución", "DISTR": "Tableros de control y Distribución",
                    "GENERAC": "Generación e Iluminación", "ILUMIN": "Generación e Iluminación",
                    "LABOR": "Laboratorio y Metrología", "METROL": "Laboratorio y Metrología", "TERMOM": "Laboratorio y Metrología",
                    "MEDIC": "Medición Proceso",
                    "HSE": "HSE", "SEGUR": "HSE", "CONTING": "HSE",
                    "CAMPAM": "Campamento",
                    "TECNOL": "Tecnología", "SISTEM": "Tecnología", "TECNO": "Tecnología",
                    "INSTRUM": "Materiales de Instrumentación",
                    "HERRAM": "Herramienta",
                    "FERRET": "Ferreteria",
                }
                for key, value in mapping.items():
                    if key in n: return value
                return None # Permitir que quede sin categoría si no hay match claro

            def normalizar_sede(nombre):
                if not nombre: return None
                n = str(nombre).strip().upper()
                if "CHICHIMENE" in n: return "Chichimene"
                if "MOCOA" in n: return "Mocoa"
                if "SIBATE" in n or "SIBATÉ" in n: return "Sibaté"
                return None

            def normalizar_uom(val):
                if not val: return "UNIDAD"
                n = str(val).strip().upper()
                validas = ["UNIDAD", "GALONES", "BARRILES", "LITROS", "METROS", "PULGADAS", "PIES"]
                for v in validas:
                    if v in n: return v
                return "UNIDAD"

            def clean_float(val):
                if val is None or val == '-': return 0.0
                try:
                    if isinstance(val, (int, float)): return float(val)
                    return float(str(val).replace(',', '.'))
                except: return 0.0

            def clean_str(val):
                if val is None: return None
                s = str(val).strip()
                if s.upper() in ["-", "N/A", "NONE", "NULL", "0", "0.0"]: return None
                return s

            for row_idx, row_data in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2): # Se añade enumerate para obtener el índice de la fila
                # 1. Saltar filas vacías o con muy pocos datos
                if not any(row_data) or len(row_data) < 21: # Expecting at least 21 columns
                    continue

                # Mapeo según el orden de 21 columnas solicitado (Tipo hasta Fecha Calibración/Garantía):
                tipo_equipo = clean_str(row_data[0]) # 0: Tipo
                item_name_raw = clean_str(row_data[3]) # 3: Descripción (Nombre)
                
                # Ignorar encabezados
                if (tipo_equipo and tipo_equipo.upper() in ["TIPO", "TIPO DE EQUIPO"]) or (item_name_raw and item_name_raw.upper() in ["DESCRIPCION DE ARTICULO", "DESCRIPCIÓN"]):
                    continue

                # 1: Fecha Creación
                fecha_creacion = timezone.now()
                if row_data[1]:
                    try:
                        if isinstance(row_data[1], datetime):
                            if timezone.is_naive(row_data[1]):
                                fecha_creacion = timezone.make_aware(row_data[1])
                            else:
                                fecha_creacion = row_data[1]
                        else: fecha_creacion = timezone.make_aware(datetime.strptime(str(row_data[1]), '%d/%m/%Y'))
                    except: pass
                
                # 2: Categoría
                cat_name_norm = normalizar_categoria(row_data[2])
                cat_obj = None
                if cat_name_norm:
                    cat_obj, _ = InventoryCategory.objects.get_or_create(name=cat_name_norm)

                # Map user's columns to row_data indices
                tech_spec_raw = clean_str(row_data[4])
                brand_raw = clean_str(row_data[5])
                serial_raw = clean_str(row_data[6])
                tag_raw = clean_str(row_data[7])
                uom_val = normalizar_uom(row_data[8])

                length_val = clean_float(row_data[9])
                width_val = clean_float(row_data[10])
                height_val = clean_float(row_data[11])
                weight_val = clean_float(row_data[12])
                
                # User's list does not have 'cost' or 'calibration_cost' explicitly.
                # Use default values for these.
                cost_val = 0.00
                calibration_cost_val = 0.00

                # User's column 13 is "stock total", column 15 is "stock disponible"
                # We will use "stock disponible" (index 15) for current_stock
                stock_val = int(clean_float(row_data[15])) if row_data[15] is not None else 0
                
                owner_val = clean_str(row_data[14]) # User's column 14 is "PROPIEDAD"
                status_val = clean_str(row_data[16]) or "OPERATIVO" # User's column 16 is "ESTADO"
                desc_val = clean_str(row_data[17]) # User's column 17 is "OBSERVACIONES"

                raw_sede = clean_str(row_data[18]) # User's column 18 is "BASE O SEDE"
                sede_name_norm = normalizar_sede(raw_sede)
                sede_obj = None
                if sede_name_norm:
                    sede_obj, _ = Ubicacion.objects.get_or_create(nombre=sede_name_norm)
                
                proj_name = clean_str(row_data[19]) # User's column 19 is "PROYECTO"
                proj_obj = None
                if proj_name:
                    proj_obj, _ = Project.objects.get_or_create(name=proj_name)

                # SKU único
                sku_val = tag_raw or serial_raw
                if not sku_val:
                    import hashlib
                    sku_seed = item_name_raw if item_name_raw else "ITEM"
                    hash_name = hashlib.md5(sku_seed.encode()).hexdigest()[:6].upper()
                    sku_val = f"AUTO_{slugify(sku_seed)[:15].upper()}_{hash_name}"
                
                # Handle "FECHA DE CALIBRACIÓN / GARANTÍA" (User's column 20)
                calib_warranty_raw = row_data[20]
                calibration_date_val = None
                warranty_val = None

                if calib_warranty_raw:
                    try:
                        # Try to parse as date first
                        if isinstance(calib_warranty_raw, datetime):
                            calibration_date_val = calib_warranty_raw.date()
                        else:
                            calibration_date_val = datetime.strptime(str(calib_warranty_raw).split(' ')[0], '%d/%m/%Y').date()
                    except (ValueError, TypeError):
                        # If not a date, treat as warranty string
                        warranty_val = clean_str(calib_warranty_raw)


                item, created = InventoryItem.objects.update_or_create(
                    sku=sku_val,
                    base=sede_obj,
                    project=proj_obj,
                    defaults={
                        'equipment_type': tipo_equipo,
                        'created_at': fecha_creacion,
                        'category': cat_obj,
                        'name': item_name_raw if item_name_raw else sku_val,
                        'description': desc_val,
                        'technical_spec': tech_spec_raw,
                        'brand': brand_raw,
                        'serial': serial_raw,
                        'tag': tag_raw,
                        'unit_of_measure': uom_val,
                        'length': length_val,
                        'width': width_val,
                        'height': height_val,
                        'weight': weight_val,
                        'cost': cost_val, # Default to 0.00 as not in user's columns
                        'calibration_cost': calibration_cost_val, # Default to 0.00 as not in user's columns
                        'owner': owner_val,
                        'status': status_val,
                        'current_stock': stock_val,
                        'calibration_date': calibration_date_val,
                        'warranty': warranty_val,
                    }
                )

                # Generar Trazabilidad (Movimiento de carga inicial)
                if created and stock_val > 0:
                    StockTransaction.objects.create(
                        item=item,
                        movement_type='ENTRY',
                        quantity=stock_val,
                        responsible=ctx['usuario'],
                        reason="Carga inicial masiva de inventario (Alimentación de datos)",
                        destination_base=sede_obj,
                        destination_project=proj_obj
                    )

                if created: creados += 1
                else: actualizados += 1

            # --- BARRIDO DE LIMPIEZA TOTAL ---
            # 1. Aseguramos que las 17 existan en la DB
            for cat_name in WHITELIST_CATEGORIES:
                InventoryCategory.objects.get_or_create(name=cat_name)

            # 2. Re-mapeamos TODOS los equipos de la base de datos a las 17 oficiales
            for item in InventoryItem.objects.all():
                nombre_correcto = normalizar_categoria(item.category.name if item.category else "")
                cat_oficial = InventoryCategory.objects.get(name=nombre_correcto)
                if item.category_id != cat_oficial.id:
                    item.category = cat_oficial
                    item.save()

            # 3. ELIMINACIÓN RADICAL: Borramos cualquier categoría que no esté en la lista blanca
            InventoryCategory.objects.exclude(name__in=WHITELIST_CATEGORIES).delete()

            # Log de Auditoría
            LogActividad.objects.create(
                admin_responsable=ctx['usuario'],
                accion="Importación Masiva Inventario",
                detalles=f"Se cargaron {creados} equipos nuevos y se actualizaron {actualizados}."
            )

            messages.success(request, f"Carga masiva finalizada: {creados} registros nuevos y {actualizados} actualizaciones. Las 17 categorías se mantienen intactas.")
        except Exception as e:
            messages.error(request, f"Error al procesar el Excel: {e}")
            
    return redirect(f"{reverse('gestion_inventario')}?tab=sede")

@login_required
def crear_item_inventario(request):
    """Vista para procesar la creación de un nuevo equipo desde el modal."""
    ctx = obtener_contexto_usuario(request)
    tiene_permiso = (ctx.get('crear') and ctx.get('inventario')) or request.user.is_superuser
    if not tiene_permiso:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': "No tienes permisos para crear equipos en el inventario."}, status=403)
        messages.error(request, "No tienes permisos para crear equipos en el inventario.")
        return redirect('gestion_inventario')

    if request.method == 'POST':
        try:
            # Creamos el item y lo asignamos a la variable 'item' para usarla en la respuesta JSON
            item = InventoryItem.objects.create(
                name=request.POST.get('name'),
                sku=request.POST.get('sku'),
                description=request.POST.get('description'),
                technical_spec=request.POST.get('technical_spec'),
                category_id=request.POST.get('category') or None, # Maneja categoría vacía
                base_id=request.POST.get('base') or None,
                current_stock=int(request.POST.get('current_stock', 0)),
                unit_price=float(request.POST.get('unit_price', 0.00)),
                cost=float(request.POST.get('cost', 0.00)),
                calibration_cost=float(request.POST.get('calibration_cost', 0.00)),
                min_stock=int(request.POST.get('min_stock', 5)),
                max_stock=int(request.POST.get('max_stock', 100)),
                unit_of_measure=request.POST.get('unit_of_measure', 'UNIDAD'),
                equipment_type=request.POST.get('equipment_type'),
                brand=request.POST.get('brand'),
                serial=request.POST.get('serial'),
                tag=request.POST.get('tag'),
                owner=request.POST.get('owner'),
                project_id=request.POST.get('project') or None,
                status=request.POST.get('status'),
                length=float(request.POST.get('length', 0.00)),
                width=float(request.POST.get('width', 0.00)),
                height=float(request.POST.get('height', 0.00)),
                weight=float(request.POST.get('weight', 0.00)),
                created_at=request.POST.get('created_at') or timezone.now()
            )
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success', 'message': f"Equipo '{item.name}' creado exitosamente.", 'item_id': item.id, 'item_name': item.name, 'item_sku': item.sku, 'item_stock': item.current_stock, 'item_base': item.base.nombre if item.base else 'N/A', 'item_project': item.project.name if item.project else 'N/A'})
            messages.success(request, f"Equipo '{request.POST.get('name')}' creado exitosamente.")
        except Exception as e:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': f"Error al crear el equipo: {e}"}, status=400)
            messages.error(request, f"Error al crear el equipo: {e}")
            
    return redirect(f"{reverse('gestion_inventario')}?tab=crear")

@login_required
def editar_equipo_inventario(request, pk):
    """Procesa la actualización de los datos de un equipo existente."""
    ctx = obtener_contexto_usuario(request)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if not (ctx.get('editar') or request.user.is_superuser):
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': "No tienes permisos para editar equipos."}, status=403)
        else: # Non-AJAX fallback
            messages.error(request, "No tienes permisos para editar equipos.")
            return redirect('gestion_inventario')

    equipo = get_object_or_404(InventoryItem, pk=pk)
    
    if request.method == 'POST':
        try:
            estado_anterior = equipo.status
            stock_anterior = equipo.current_stock

            def clean_num(val, is_int=False):
                if not val or str(val).strip() == '': return 0
                cleaned = str(val).replace(',', '.')
                return int(float(cleaned)) if is_int else float(cleaned)

            with transaction.atomic():
                equipo.name = request.POST.get('name')
                equipo.sku = request.POST.get('sku')
                equipo.description = request.POST.get('description')
                equipo.technical_spec = request.POST.get('technical_spec')
                
                # Permitir limpiar campos (Set to None) si se envía vacío
                def get_id_or_none(key):
                    val = request.POST.get(key)
                    return int(val) if val and val.isdigit() else None

                equipo.category_id = get_id_or_none('category')
                equipo.base_id = get_id_or_none('base')
                equipo.project_id = get_id_or_none('project')
                equipo.supplier_id = get_id_or_none('supplier')

                equipo.unit_price = clean_num(request.POST.get('unit_price'))
                equipo.cost = clean_num(request.POST.get('cost'))
                equipo.calibration_cost = clean_num(request.POST.get('calibration_cost'))
                equipo.min_stock = clean_num(request.POST.get('min_stock'), True)
                equipo.max_stock = clean_num(request.POST.get('max_stock'), True)
                
                equipo.unit_of_measure = request.POST.get('unit_of_measure')
                equipo.equipment_type = request.POST.get('equipment_type')
                equipo.brand = request.POST.get('brand')
                equipo.serial = request.POST.get('serial')
                equipo.tag = request.POST.get('tag')
                equipo.owner = request.POST.get('owner') # Corresponde a "Propiedad"
                equipo.rented_quantity = clean_num(request.POST.get('rented_quantity'), True)
                equipo.manufacturing_quantity = clean_num(request.POST.get('manufacturing_quantity'), True)
                equipo.ods_order = request.POST.get('ods_order')
                
                # Manejo de fechas para evitar errores de formato vacío
                equipo.calibration_date = request.POST.get('calibration_date') if request.POST.get('calibration_date') else None
                equipo.maintenance_date = request.POST.get('maintenance_date') if request.POST.get('maintenance_date') else None
                equipo.warranty = request.POST.get('warranty')
                
                nuevo_stock_raw = request.POST.get('current_stock')
                if nuevo_stock_raw is not None and nuevo_stock_raw != '':
                    nuevo_stock = int(clean_num(nuevo_stock_raw, True))
                    if nuevo_stock != stock_anterior:
                        equipo.current_stock = nuevo_stock
                        diff = nuevo_stock - stock_anterior
                        StockTransaction.objects.create(
                            item=equipo,
                            movement_type='ADJUST_IN' if diff > 0 else 'ADJUST_OUT',
                            quantity=abs(diff),
                            responsible=ctx['usuario'],
                            reason=f"Ajuste manual detectado en edición (Stock anterior: {stock_anterior})"
                        )

                nuevo_estado = request.POST.get('status')
                equipo.status = nuevo_estado
                equipo.length = clean_num(request.POST.get('length'))
                equipo.width = clean_num(request.POST.get('width'))
                equipo.height = clean_num(request.POST.get('height'))
                equipo.weight = clean_num(request.POST.get('weight'))
                
                if request.POST.get('created_at'):
                    equipo.created_at = request.POST.get('created_at')

                # LÓGICA AUTOMÁTICA: Si el estado cambia a 'VENDIDO' o 'DADO DE BAJA', 
                # se liquida el stock actual y se registra la transacción de salida.
                if nuevo_estado in ['VENDIDO', 'DADO DE BAJA'] and estado_anterior not in ['VENDIDO', 'DADO DE BAJA']:
                    if equipo.current_stock > 0:
                        stock_liquidado = equipo.current_stock
                        equipo.current_stock = 0
                        
                        StockTransaction.objects.create(
                            item=equipo,
                            movement_type='EXIT',
                            quantity=stock_liquidado,
                            responsible=ctx['usuario'],
                            reason=f"Ajuste automático por cambio de estado a {nuevo_estado}"
                        )
                        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                            # For AJAX, we don't use Django messages, but return JSON
                            pass
                        else:
                            messages.info(request, f"Stock de '{equipo.name}' ajustado a 0 automáticamente.")
                            
                equipo.save()
                
                # Log de Auditoría
                LogActividad.objects.create(
                    usuario_afectado=ctx['usuario'], # Se registra quién realizó el cambio
                    admin_responsable=ctx['usuario'],
                    accion="Edición de Equipo",
                    detalles=f"Actualización técnica del equipo SKU: {equipo.sku} - {equipo.name}"
                )
            if is_ajax:
                return JsonResponse({'status': 'success', 'message': f"¡El cambio fue exitoso! Equipo '{equipo.name}' actualizado correctamente."})
            messages.success(request, f"¡El cambio fue exitoso! Equipo '{equipo.name}' actualizado correctamente.")
        except Exception as e:
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': f"Error al actualizar el equipo: {e}"}, status=400)
            messages.error(request, f"Error al actualizar el equipo: {e}")
            
    return redirect(f"{reverse('gestion_inventario')}?tab=sede")

@login_required
def crear_proyecto(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        observations = request.POST.get('observations')
        if name:
            try:
                project = Project.objects.create(name=name, observations=observations)
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'status': 'success', 'message': f"Proyecto '{name}' creado correctamente.", 'project_id': project.id, 'project_name': project.name, 'project_observations': project.observations})
                messages.success(request, f"Proyecto '{name}' creado correctamente.")
            except Exception as e:
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'status': 'error', 'message': f"Error al crear el proyecto: {e}"}, status=400)
                messages.error(request, f"Error al crear el proyecto: {e}")
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': "El nombre del proyecto no puede estar vacío."}, status=400)
            messages.error(request, "El nombre del proyecto no puede estar vacío.")
    return redirect(f"{reverse('gestion_inventario')}?tab=crear")

@login_required
def editar_proyecto(request, pk):
    ctx = obtener_contexto_usuario(request)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if not (ctx.get('editar') or request.user.is_superuser):
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': "No tienes permisos para editar proyectos."}, status=403)
        messages.error(request, "No tienes permisos para editar proyectos.")
        return redirect('gestion_inventario')

    proyecto = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        try:
            proyecto.name = request.POST.get('name')
            proyecto.observations = request.POST.get('observations')
            proyecto.save()
            if is_ajax:
                return JsonResponse({'status': 'success', 'message': f"Proyecto '{proyecto.name}' actualizado.", 'project_id': proyecto.id, 'project_name': proyecto.name, 'project_observations': proyecto.observations})
            messages.success(request, f"Proyecto '{proyecto.name}' actualizado.")
        except Exception as e:
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': f"Error al actualizar el proyecto: {e}"}, status=400)
            messages.error(request, f"Error al actualizar el proyecto: {e}")
    return redirect(f"{reverse('gestion_inventario')}?tab=crear")

@login_required
def eliminar_proyecto(request, pk):
    ctx = obtener_contexto_usuario(request)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if not (ctx.get('eliminar') or request.user.is_superuser):
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': "No tienes permisos para eliminar proyectos."}, status=403)
        messages.error(request, "No tienes permisos para eliminar proyectos.")
        return redirect('gestion_inventario')

    proyecto = get_object_or_404(Project, pk=pk)
    nombre = proyecto.name
    try:
        proyecto.delete()
        if is_ajax:
            return JsonResponse({'status': 'success', 'message': f"Proyecto '{nombre}' eliminado.", 'project_id': pk})
        messages.success(request, f"Proyecto '{nombre}' eliminado.")
    except Exception as e:
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': f"Error al eliminar el proyecto: {e}"}, status=400)
        messages.error(request, f"Error al eliminar el proyecto: {e}")
    return redirect(f"{reverse('gestion_inventario')}?tab=crear")

@login_required
def eliminar_equipo_inventario(request, pk):
    """Elimina un equipo del inventario de forma permanente."""
    ctx = obtener_contexto_usuario(request)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if not (ctx.get('eliminar') or request.user.is_superuser):
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': "No tienes permisos para eliminar equipos."}, status=403)
        messages.error(request, "No tienes permisos para eliminar equipos.")
        return redirect('gestion_inventario')

    equipo = get_object_or_404(InventoryItem, pk=pk)
    nombre = equipo.name
    sku = equipo.sku
    
    try:
        # Registro en el log de auditoría antes de eliminar
        LogActividad.objects.create(
            admin_responsable=ctx['usuario'],
            accion="Eliminación de Equipo",
            detalles=f"Eliminación permanente: {nombre} (SKU: {sku}) - Sede: {equipo.base.nombre if equipo.base else 'N/A'}"
        )
        equipo.delete()
        if is_ajax:
            return JsonResponse({'status': 'success', 'message': f"Equipo '{nombre}' eliminado correctamente.", 'item_id': pk})
        messages.success(request, f"Equipo '{nombre}' eliminado correctamente.")
    except Exception as e:
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': f"Error al intentar eliminar el equipo: {e}"}, status=400)
        messages.error(request, f"Error al intentar eliminar el equipo: {e}")
        
    return redirect('gestion_inventario')

@login_required
def new_record_stock_movement(request):
    """
    Registra una entrada o salida de stock de forma atómica.
    """
    ctx = obtener_contexto_usuario(request)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if not ctx['inventario']: # Verifica el permiso de inventario
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': "No tienes permisos para registrar movimientos de inventario."}, status=403)
        messages.error(request, "No tienes permisos para registrar movimientos de inventario.")
        return redirect('gestion_inventario')

    if request.method == 'POST':
        item_id = request.POST.get('item_id')
        movement_type = request.POST.get('movement_type') # 'ENTRY' o 'EXIT'
        quantity_str = request.POST.get('quantity')
        reason = request.POST.get('reason')
        project_id = request.POST.get('project_id')
        dest_base_id = request.POST.get('dest_base_id')

        try:
            quantity = int(quantity_str)
            if quantity <= 0:
                raise ValueError("La cantidad debe ser mayor a cero.")
            
            with transaction.atomic():
                item = get_object_or_404(InventoryItem, id=item_id)
                project = Project.objects.filter(id=project_id).first() if project_id else None
                dest_base = Ubicacion.objects.filter(idUbicacion=dest_base_id).first() if dest_base_id else None

                # Lógica de Control: Determinar si el movimiento reduce el stock físico
                es_reduccion = movement_type in ['EXIT', 'ADJUST_OUT']
                
                if es_reduccion:
                    if item.current_stock < quantity:
                        if is_ajax:
                            return JsonResponse({'status': 'error', 'message': f"Error: Stock insuficiente para {item.name}. Disponible: {item.current_stock}"}, status=400)
                        messages.error(request, f"Error: Stock insuficiente para {item.name}. Disponible: {item.current_stock}")
                        return redirect(f"{reverse('gestion_inventario')}?tab=sede")
                    item.current_stock -= quantity
                else: # Incremento: ENTRY, RETURN, ADJUST_IN
                    if item.current_stock + quantity > item.max_stock:
                        if not is_ajax:
                            messages.warning(request, f"Atención: El stock resultante de {item.name} ({item.current_stock + quantity}) supera el máximo definido ({item.max_stock}).")
                    item.current_stock += quantity
                
                # Lógica de estados automática para Retornos y Salidas a Proyecto
                if movement_type == 'EXIT' and project:
                    item.status = 'EN PROYECTO'
                elif movement_type == 'RETURN':
                    item.status = 'OPERATIVO' # Vuelve a estar Disponible
                    if dest_base:
                        item.base = dest_base # Actualizar ubicación física del equipo
                
                item.save()

                trans = StockTransaction.objects.create(
                    item=item,
                    movement_type=movement_type,
                    quantity=quantity,
                    responsible=ctx['usuario'],
                    reason=reason,
                    # Registrar el proyecto relacionado en los campos de auditoría
                    origin_project=project if movement_type == 'RETURN' else None,
                    destination_project=project if movement_type == 'EXIT' else None,
                    destination_base=dest_base if movement_type == 'RETURN' else None
                )
                
                # Log de Auditoría para Control Interno
                LogActividad.objects.create(
                    admin_responsable=ctx['usuario'],
                    accion="Movimiento Stock Manual",
                    detalles=f"{trans.get_movement_type_display()} de {quantity} unidades para {item.name} (SKU: {item.sku})."
                )
                
                if is_ajax:
                    return JsonResponse({'status': 'success', 'message': f"Movimiento registrado exitosamente para {item.name}."})
                messages.success(request, f"Movimiento registrado exitosamente para {item.name}.")
        except (ValueError, TypeError) as e:
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': f"La cantidad ingresada no es válida o hay un problema con los datos: {e}"}, status=400)
            messages.error(request, f"La cantidad ingresada no es válida o hay un problema con los datos: {e}")
        except Exception as e:
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': f"Ocurrió un error inesperado al registrar el movimiento: {e}"}, status=400)
            messages.error(request, f"Ocurrió un error inesperado al registrar el movimiento: {e}")
    
    return redirect(f"{reverse('gestion_inventario')}?tab=sede")

@login_required
def transferir_item_inventario(request):
    """
    Mueve stock de un equipo de una base a otra. 
    Si el equipo no existe en la base destino (mismo SKU), se crea automáticamente.
    """
    ctx = obtener_contexto_usuario(request)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if not ctx['inventario']:
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': "No tienes permisos para realizar traslados."}, status=403)
        messages.error(request, "No tienes permisos para realizar traslados.")
        return redirect('gestion_inventario')

    if request.method == 'POST':
        item_id = request.POST.get('item_id')
        dest_base_id = request.POST.get('dest_base_id')
        dest_project_id = request.POST.get('dest_project_id')
        quantity = int(request.POST.get('quantity', 0))
        reason = request.POST.get('reason', 'Traslado de equipo')

        try:
            with transaction.atomic():
                item_origen = get_object_or_404(InventoryItem, id=item_id)
                # Se permite que la base o el proyecto sean opcionales, pero al menos uno debe existir
                base_destino = Ubicacion.objects.filter(idUbicacion=dest_base_id).first() if dest_base_id else None
                proyecto_destino = Project.objects.filter(id=dest_project_id).first() if dest_project_id else None

                if not base_destino and not proyecto_destino:
                    if is_ajax:
                        return JsonResponse({'status': 'error', 'message': "Debe seleccionar al menos una Base o un Proyecto de destino."}, status=400)
                    messages.error(request, "Debe seleccionar al menos una Base o un Proyecto de destino.")
                    return redirect(f"{reverse('gestion_inventario')}?tab=sede")

                if item_origen.base == base_destino and item_origen.project == proyecto_destino:
                    if is_ajax:
                        return JsonResponse({'status': 'error', 'message': "La ubicación y proyecto de destino no pueden ser idénticos al origen."}, status=400)
                    messages.error(request, "La ubicación y proyecto de destino no pueden ser idénticos al origen.")
                    return redirect(f"{reverse('gestion_inventario')}?tab=sede")

                if item_origen.current_stock < quantity:
                    if is_ajax:
                        return JsonResponse({'status': 'error', 'message': f"Stock insuficiente en {item_origen.base.nombre}. Disponible: {item_origen.current_stock}"}, status=400)
                    messages.error(request, f"Stock insuficiente en {item_origen.base.nombre}. Disponible: {item_origen.current_stock}")
                    return redirect(f"{reverse('gestion_inventario')}?tab=sede")

                # 1. Restar del origen
                item_origen.current_stock -= quantity
                item_origen.save()

                # 2. Sumar al destino (buscar por SKU en esa base)
                item_destino, created = InventoryItem.objects.get_or_create(
                    sku=item_origen.sku,
                    base=base_destino,
                    project=proyecto_destino,
                    defaults={
                        'name': item_origen.name,
                        'category': item_origen.category,
                        'unit_price': item_origen.unit_price,
                        'unit_of_measure': item_origen.unit_of_measure,
                        'min_stock': item_origen.min_stock,
                        'max_stock': item_origen.max_stock,
                    }
                )
                item_destino.current_stock += quantity
                item_destino.save()

                # 3. Registrar transacción
                StockTransaction.objects.create(
                    item=item_origen,
                    movement_type='TRANSFER',
                    quantity=quantity,
                    responsible=ctx['usuario'],
                    reason=reason,
                    origin_base=item_origen.base,
                    destination_base=base_destino,
                    origin_project=item_origen.project,
                    destination_project=proyecto_destino
                )
                if is_ajax:
                    return JsonResponse({'status': 'success', 'message': f"Traslado de {quantity} unidades de '{item_origen.name}' de {item_origen.base.nombre} a {base_destino.nombre} exitoso."})
                messages.success(request, f"Traslado de {quantity} unidades de '{item_origen.name}' de {item_origen.base.nombre} a {base_destino.nombre} exitoso.")
        except Exception as e:
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': f"Error en el traslado: {e}"}, status=400)
            messages.error(request, f"Error en el traslado: {e}")
            
    return redirect(f"{reverse('gestion_inventario')}?tab=sede")

@login_required
def solicitar_transferencia(request):
    """Permite a los Supervisores crear una petición de traslado."""
    ctx = obtener_contexto_usuario(request)
    
    # Verificación robusta del rol
    if not ctx['is_admin'] and not ctx['is_supervisor']:
        messages.error(request, "Solo los supervisores o administradores pueden realizar solicitudes de traslado.")
        return redirect(f"{reverse('gestion_inventario')}?tab=sede")

    if request.method == 'POST':
        dest_base_id = request.POST.get('dest_base_id')
        dest_project_id = request.POST.get('dest_project_id')
        reason = request.POST.get('reason')

        items_data = request.POST.get('items_json')
        if not items_data:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': "No hay equipos en la lista."}, status=400)
            messages.error(request, "No hay equipos seleccionados.")
            return redirect(f"{reverse('gestion_inventario')}?tab=requests")

        try:
            items_list = json.loads(items_data)
        except:
            return JsonResponse({'status': 'error', 'message': "Error en formato de datos."}, status=400)

        if not dest_base_id and not dest_project_id:
            messages.error(request, "Debe seleccionar una Base o un Proyecto de destino para la solicitud.")
            return redirect(f"{reverse('gestion_inventario')}?tab=sede")

        batch_id = str(uuid.uuid4())[:8].upper()
        for entry in items_list:
            item = get_object_or_404(InventoryItem, id=entry['id'])
            TransferRequest.objects.create(
                item=item, quantity=int(entry['qty']),
                destination_base_id=dest_base_id or None,
                destination_project_id=dest_project_id or None,
                supervisor=ctx['usuario'], reason=reason,
                batch_id=batch_id # Siempre guardamos el batch_id para tratarlo como solicitud única
            )
        
        # Notificación por correo al Administrador
        try:
            send_mail(
                'Petrocentro - Nueva Solicitud de Traslado de Equipos',
                f'El usuario {ctx["usuario"].nombre} ha realizado una nueva solicitud de traslado de equipos.\n\nMotivo: {reason}\n\nPor favor, ingrese al sistema para aprobar o rechazar esta solicitud.',
                settings.EMAIL_HOST_USER,
                [settings.EMAIL_HOST_USER], # Se envía al correo del administrador
                fail_silently=True
            )
        except:
            pass
            
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'ok'})
            
        messages.success(request, "Tu solicitud se envió exitosamente.")
    return redirect(f"{reverse('gestion_inventario')}?tab=requests")

@login_required
def procesar_transferencia(request, request_id):
    """Permite al Administrador aprobar o rechazar la solicitud."""
    ctx = obtener_contexto_usuario(request)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if not ctx['is_admin']:
        messages.error(request, "No tienes permisos para procesar solicitudes.")
        return redirect(f"{reverse('gestion_inventario')}?tab=requests")

    # Obtenemos la solicitud base
    base_req = get_object_or_404(TransferRequest, id=request_id)
    # Si tiene batch_id, procesamos a todos los "hermanos"
    if base_req.batch_id:
        requests_to_process = TransferRequest.objects.filter(batch_id=base_req.batch_id, status='PENDING')
    else:
        requests_to_process = [base_req]

    action = request.POST.get('action') 
    rejection_reason = request.POST.get('rejection_reason', '')

    if action == 'reject':
        for r in requests_to_process:
            r.status = 'REJECTED'
            r.processed_by = ctx['usuario']
            r.processed_at = timezone.now()
            r.admin_notes = rejection_reason if rejection_reason.strip() else "Rechazado por el administrador."
            r.save()

        # Notificación automática al supervisor
        try:
            send_mail(
                'Solicitud de Traslado RECHAZADA - Petrocentro',
                f'Hola {base_req.supervisor.nombre},\n\nTu solicitud de traslado ha sido rechazada.\n\nMotivo del rechazo:\n{rejection_reason if rejection_reason.strip() else "Sin comentarios adicionales."}\n\nAtentamente,\nControl de Inventarios Petrocentro.',
                settings.EMAIL_HOST_USER,
                [base_req.supervisor.correo],
                fail_silently=True
            )
        except: pass

        messages.warning(request, "La solicitud ha sido rechazada.")
        return redirect(f"{reverse('gestion_inventario')}?tab=requests")

    if action == 'approve':
        try:
            with transaction.atomic():
                for r in requests_to_process:
                    item_origen = r.item
                    if item_origen.current_stock < r.quantity:
                        raise ValueError(f"Stock insuficiente para {item_origen.name}")

                    item_origen.current_stock -= r.quantity
                    item_origen.save()

                    StockTransaction.objects.create(
                        item=item_origen, movement_type='TRANSFER', quantity=r.quantity,
                        responsible=ctx['usuario'], reason=f"Lote aprobado: {r.reason}",
                        origin_base=item_origen.base, destination_base=r.destination_base,
                        origin_project=item_origen.project, destination_project=r.destination_project
                    )
                    
                    r.status = 'SENT'
                    r.processed_by = ctx['usuario']
                    r.processed_at = timezone.now()
                    r.save()

                # Log de Auditoría para Control Interno
                LogActividad.objects.create(
                    admin_responsable=ctx['usuario'],
                    accion="Aprobación de Traslado",
                    detalles=f"Traslado ID {base_req.id} aprobado (Lote: {base_req.batch_id or 'N/A'})."
                )

                # 1. Preparar datos para el PDF y el link
                id_para_link = base_req.batch_id if base_req.batch_id else str(base_req.id)
                consecutivo = base_req.batch_id[:8] if base_req.batch_id else f"REQ-{base_req.id}"
                url_pdf = reverse('descargar_requisicion_pdf', kwargs={'batch_id': id_para_link})

                # 2. Generar PDF en memoria para enviar por correo al supervisor
                msg = "" # Pre-definir mensaje
                try: # Propagar error al exterior para que no salga "Exitoso" si falla el mail
                    template = get_template('dash/pdf_requisicion.html')
                    pdf_ctx = {
                        'requests': requests_to_process,
                        'base_req': base_req,
                        'fecha_aprobacion': timezone.now(),
                        'consecutivo': consecutivo,
                        'es_compra': True,
                    }
                    html = template.render(pdf_ctx)
                    result = io.BytesIO()
                    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
                    
                    if not pdf.err:
                        # Enviar Email con Adjunto
                        email_supervisor = EmailMessage(
                            f'Requisición Aprobada GCL-FR-04: {consecutivo}',
                            f'Hola {base_req.supervisor.nombre},\n\nTu solicitud de traslado ha sido aprobada. Adjunto encontrarás el documento oficial de requisición para el despacho de los equipos.\n\nAtentamente,\nControl de Inventarios Petrocentro.',
                            settings.EMAIL_HOST_USER,
                            [base_req.supervisor.correo],
                        )
                        email_supervisor.attach(f'GCL-FR-04_{consecutivo}.pdf', result.getvalue(), 'application/pdf')
                        email_supervisor.send(fail_silently=True)
                except Exception as e_pdf:
                    raise Exception(f"Fallo en PDF/Correo: {e_pdf}")
                
                # 3. Mostrar mensaje de éxito con botón de descarga para el Admin
                msg = f'¡Aprobación exitosa! El stock ha sido descontado. <a href="{url_pdf}" id="auto-download-pdf" target="_blank" class="btn btn-sm btn-danger fw-bold ms-2 shadow-sm"><i class="fas fa-file-pdf me-1"></i>DESCARGAR REQUISICIÓN GCL-FR-04</a>'
                messages.success(request, mark_safe(msg), extra_tags='pdf-ready')
                    
        except Exception as e:
            if is_ajax: return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
            messages.error(request, f"Error al procesar: {e}")
            
    if is_ajax:
        # Si es AJAX, devolvemos el mensaje de éxito para que SweetAlert lo renderice bien
        return JsonResponse({'status': 'success', 'message': msg})
    return redirect(f"{reverse('gestion_inventario')}?tab=requests")

@login_required
def descargar_requisicion_pdf(request, batch_id):
    """Genera el PDF oficial de Requisición de Compra o Servicio GCL-FR-04 basado en un lote de solicitud."""
    requests_qs = TransferRequest.objects.filter(batch_id=batch_id).select_related('item', 'supervisor', 'processed_by', 'destination_base', 'destination_project')
    
    if not requests_qs.exists():
        # Fallback para solicitudes individuales antiguas o por ID directo
        if batch_id.isdigit():
            requests_qs = TransferRequest.objects.filter(id=batch_id).select_related('item', 'supervisor', 'processed_by', 'destination_base', 'destination_project')

    if not requests_qs.exists():
        return HttpResponse("Documento de requisición no encontrado.", status=404)

    base_req = requests_qs.first()
    
    context = {
        'requests': requests_qs,
        'base_req': base_req,
        'fecha_aprobacion': base_req.processed_at or timezone.now(),
        'consecutivo': base_req.batch_id[:8] if base_req.batch_id else f"REQ-{base_req.id}",
        'es_compra': True, # Por defecto Compra para equipos de inventario
    }

    # Renderizado del template PDF
    template = get_template('dash/pdf_requisicion.html')
    html = template.render(context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="GCL-FR-04_REQ_{context["consecutivo"]}.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error técnico al generar el PDF.', status=500)
    return response

@login_required
def recibir_transferencia(request, request_id):
    """Permite al supervisor confirmar la llegada del material y registrar novedades."""
    ctx = obtener_contexto_usuario(request)
    t_request = get_object_or_404(TransferRequest, id=request_id)
    
    # Si tiene batch_id, operamos sobre todo el lote para facilitar el trabajo del supervisor
    if t_request.batch_id:
        requests_to_process = TransferRequest.objects.filter(batch_id=t_request.batch_id, status='SENT')
    else:
        requests_to_process = [t_request]

    # Solo el supervisor de la base destino (o admin) puede confirmar recepción
    if not request.user.is_superuser and (not ctx['is_supervisor'] or ctx['empleado'].id_ubicacion != t_request.destination_base):
        messages.error(request, "No tienes autorización para recibir este material.")
        return redirect(f"{reverse('gestion_inventario')}?tab=requests")

    if request.method == 'POST':
        incidents = request.POST.get('incidents', '')

        try:
            with transaction.atomic():
                for req in requests_to_process:
                    # En recepción por lote, el stock se suma automáticamente según lo enviado
                    item_origen = req.item
                    item_destino, created = InventoryItem.objects.get_or_create(
                        sku=item_origen.sku,
                        base=req.destination_base,
                        project=req.destination_project,
                        defaults={
                            'name': item_origen.name,
                            'category': item_origen.category,
                            'unit_price': item_origen.unit_price,
                            'unit_of_measure': item_origen.unit_of_measure,
                            'equipment_type': item_origen.equipment_type,
                        }
                    )

                    item_destino.current_stock += req.quantity
                    item_destino.save()

                    req.status = 'COMPLETED'
                    req.received_quantity = req.quantity
                    req.incidents = incidents
                    req.received_at = timezone.now()
                    req.save()

                    StockTransaction.objects.create(
                        item=item_destino, movement_type='ENTRY', quantity=req.quantity,
                        responsible=ctx['usuario'], reason=f"Recibo de lote: {req.batch_id}. Obs: {incidents}"
                    )

                messages.success(request, f"Recepción de lote '{t_request.batch_id or t_request.id}' confirmada exitosamente.")
        except Exception as e:
            messages.error(request, f"Error al registrar recepción: {e}")

    return redirect(f"{reverse('gestion_inventario')}?tab=requests")
